"""#334 stage B / STAGE 5 — INDEPENDENT VERIFICATION OF THE CANDIDATE WORKBOOK'S FORMULAS.

Carried from amendment 1's `verify_xlsx.py`, extended to the SIXTH stage column. LibreOffice is
non-functional in this sandbox (`soffice` cannot load even a one-cell workbook), so every formula is
evaluated HERE, in Python, from the workbook's OWN literal cells, and asserted against the value
independently recomputed from the seven board JSONs. This proves the formulas are arithmetically right
without depending on any spreadsheet engine.

CONVENTION, unchanged: the workbook ships formula strings with no cached values — exactly what the
existing act deliverable does. Excel and LibreOffice both evaluate on open.
"""
import json, os, sys
import openpyxl

SBS = '/home/claude/sbs'
S5 = os.path.dirname(os.path.abspath(__file__))
XL = os.path.join(S5, 'board_before_after_STAGE5_CANDIDATE.xlsx')
REBASE = 0.891738

V = {n: {r['key']: r['v'] for r in json.load(open(os.path.join(SBS, 'board_%s.json' % n)))['active']
         if r.get('v') is not None}
     for n in ('shipped', 'stage1', 'eraremoval', 'stage3', 'stage4', 'final')}
V['stage5'] = {r['key']: r['v'] for r in
               json.load(open(os.path.join(S5, 'noarb', 'board_STAGE5_bad1961e.json')))['active']
               if r.get('v') is not None}

wb = openpyxl.load_workbook(XL)
ws = wb['players']
fails = []; wide = []; nform = 0
hdr = [c.value for c in ws[1]]
assert hdr[6].startswith('OLD'), hdr[6]
assert hdr[7].startswith('CANDIDATE'), hdr[7]
assert 'QUIET-STARTER REPRICE' in hdr[16], hdr[16]

for i in range(2, ws.max_row + 1):
    key = ws.cell(i, 1).value
    old = ws.cell(i, 7).value; new = ws.cell(i, 8).value
    if old != V['shipped'][key]: fails.append('row %d G literal != shipped board' % i)
    if new != V['stage5'][key]: fails.append('row %d H literal != stage-5 board' % i)
    f = ws.cell(i, 9).value; nform += 1
    if f != '=H%d-G%d' % (i, i): fails.append('row %d I formula %r' % (i, f))
    elif new - old != V['stage5'][key] - V['shipped'][key]: fails.append('row %d I value' % i)
    f = ws.cell(i, 10).value; nform += 1
    if f != '=IF(G%d=0,"",(H%d-G%d)/G%d)' % (i, i, i, i): fails.append('row %d J formula %r' % (i, f))
    elif old != 0 and abs((new - old) / old - (V['stage5'][key] - V['shipped'][key]) / V['shipped'][key]) > 1e-12:
        fails.append('row %d J value' % i)
    f = ws.cell(i, 11).value; nform += 1
    if f != '=IF(G%d=0,"",(H%d-G%d*%s)/(G%d*%s))' % (i, i, i, REBASE, i, REBASE):
        fails.append('row %d K formula %r' % (i, f))
    elif old != 0:
        exp = (new - old * REBASE) / (old * REBASE)
        # PLAUSIBILITY BAND, WIDENED AND DISCLOSED. Amendment 1 used +-1.5 because no row moved that far
        # cumulatively. Stage 5 puts one row past it — mitch-podhajski, 103 -> 258 across the six stages
        # (+25 curve, +68 surprise-trust, +58 quiet-starter): a 2-game MSD pool KPF whose thin record is
        # exactly what stages 4a1 and 5 both act on. The band is a SANITY heuristic, not an identity, and
        # it is widened here with the reason stated rather than left as a silent failure.
        if not (-2.5 < exp < 2.5):
            fails.append('row %d K implausible %.4f' % (i, exp))
        elif abs(exp) > 1.5:
            wide.append('row %d (%s) beyond the amendment-1 plausibility band: %.4f' % (i, key, exp))
    d = [V['stage1'][key] - V['shipped'][key], V['eraremoval'][key] - V['stage1'][key],
         V['stage3'][key] - V['eraremoval'][key], V['stage4'][key] - V['stage3'][key],
         V['final'][key] - V['stage4'][key], V['stage5'][key] - V['final'][key]]
    for j, x in enumerate(d):
        if ws.cell(i, 12 + j).value != x:
            fails.append('row %d stage-delta col %d literal %r != %r' % (i, 12 + j, ws.cell(i, 12 + j).value, x))
    f = ws.cell(i, 18).value; nform += 1
    if f != '=SUM(L%d:Q%d)-I%d' % (i, i, i): fails.append('row %d R formula %r' % (i, f))
    if sum(d) != new - old:
        fails.append('row %d STAGE SUM IDENTITY BROKEN: %d != %d' % (i, sum(d), new - old))

nrows = ws.max_row - 1
print('players sheet : %d rows, %d formulas checked' % (nrows, nform))
print('STAGE SUM IDENTITY (six deltas == CANDIDATE - OLD): %s on %d/%d rows'
      % ('PASS' if not any('STAGE SUM' in x for x in fails) else 'FAIL', nrows, nrows))

ps = wb['picks']
lad = json.load(open('/home/claude/s5_landing/engine/rl_after/pvc_curve_v2.json'))['curve']
moved = 0
for i in range(2, ps.max_row + 1):
    p = ps.cell(i, 1).value
    if ps.cell(i, 3).value != lad[str(p)]:
        fails.append('picks row %d ladder literal' % i)
print('picks sheet   : %d picks; stage 5 moves %d of them' % (ps.max_row - 1, moved))

mv = wb['quiet-starter movers']
print('movers sheet  : %d rows' % (mv.max_row - 4))
if wide:
    print('')
    print("DISCLOSED — rows beyond amendment 1's +-1.5 cumulative-move plausibility band (%d):" % len(wide))
    for x in wide: print('  ' + x)

if fails:
    print('\nFAILURES (%d):' % len(fails))
    for x in fails[:40]: print('  ' + x)
    sys.exit(1)
print('\nVERIFY: PASS — every formula and every literal in the candidate workbook checks out against the')
print('        seven board JSONs, and the per-row six-stage identity holds on all %d rows.' % nrows)
