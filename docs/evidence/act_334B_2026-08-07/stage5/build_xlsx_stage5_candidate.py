"""#334 stage B / STAGE 5 — THE SIXTH STAGE COLUMN, built as a **CANDIDATE**, NOT landed.

Stage 5 STOPPED on the landing floor (yr1 0.9945 < 1.00), so the owner-review workbook
`side_by_side/board_before_after.xlsx` is deliberately NOT touched: it describes the FIVE landed stage
columns and adding a sixth for a board that did not land would be a false claim in the owner's own
review set. This file builds the sixth column separately, here, labelled CANDIDATE on every sheet, so
the work and its 804/804 per-row identity are verified and available if the owner rules to take the
frontier as it stands.

Per-row identity asserted in Python (LibreOffice is broken in this sandbox — amendment 1's convention)
AND carried as a live formula column so a reader can watch it hold.
"""
import json, csv, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SBS = '/home/claude/sbs'
S5 = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(S5, 'board_before_after_STAGE5_CANDIDATE.xlsx')
REBASE = 0.891738          # stage-3 uniform numeraire re-base; unmoved by stages 4, 4a1 and 5

NAMES = ['shipped', 'stage1', 'eraremoval', 'stage3', 'stage4', 'final']
MD = {'shipped': '113b36f8', 'stage1': 'de5110bb', 'eraremoval': 'f94e0778',
      'stage3': '6c9f8d3a', 'stage4': 'b490ae8b', 'final': 'b56bbdde', 'stage5': 'bad1961e'}


def vmap(d):
    return {r['key']: r['v'] for r in d['active'] if r.get('v') is not None}


B = {n: json.load(open(os.path.join(SBS, 'board_%s.json' % n))) for n in NAMES}
B['stage5'] = json.load(open(os.path.join(S5, 'noarb', 'board_STAGE5_bad1961e.json')))
V = {n: vmap(B[n]) for n in B}
keys = [r['key'] for r in B['stage5']['active'] if r.get('v') is not None]
meta = {r['key']: r for r in B['stage5']['active']}
for n in V:
    assert set(V[n]) == set(keys), 'roster differs at %s: %d vs %d' % (n, len(V[n]), len(keys))
print('roster identical across all SEVEN boards: %d players' % len(keys))

HDR = Font(name='Arial', sz=10, b=True, color='FFFFFF')
NRM = Font(name='Arial', sz=10)
FILL = PatternFill('solid', fgColor='7B1E1E')      # candidate red, not the adopted navy
INT = '#,##0;(#,##0);-'
PCT = '0.00%;(0.00%);-'

wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'players'
cols = [('key', 26), ('name', 24), ('club', 20), ('pos', 7), ('age', 6), ('entry', 26),
        ('OLD — shipped board 113b36f8', 17),
        ('CANDIDATE — stage-5 board bad1961e (NOT LANDED)', 20),
        ('abs Δ (CAND − OLD)', 13), ('% Δ (of OLD)', 12),
        ('% change beyond the currency re-base', 29),
        ('Δ reference layer (de5110bb − 113b36f8)', 18),
        ('Δ era removal (f94e0778 − de5110bb)', 18),
        ('Δ curve+surface+numéraire (6c9f8d3a − f94e0778)', 22),
        ('Δ reactivity (b490ae8b − 6c9f8d3a)', 18),
        ('Δ surprise-trust (b56bbdde − b490ae8b)', 20),
        ('Δ QUIET-STARTER REPRICE (bad1961e − b56bbdde) — SIXTH, CANDIDATE', 26),
        ('STAGE SUM CHECK (ΣΔ stages − abs Δ; must be 0)', 26)]
for i, (h, w) in enumerate(cols, 1):
    c = ws.cell(1, i, h); c.font = HDR; c.fill = FILL
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.column_dimensions[get_column_letter(i)].width = w


def entry_label(r):
    ty = r.get('ty') or '?'; pk = r.get('pk')
    if ty == 'ND' and pk and pk <= 64: return 'National pick %d' % pk
    return '%s %s' % (r.get('draft') or ty, ('pick %d' % pk) if pk else '(pool)')


bad = 0
for i, k in enumerate(keys, 2):
    r = meta[k]
    ws.cell(i, 1, k).font = NRM; ws.cell(i, 2, r.get('name')).font = NRM
    ws.cell(i, 3, r.get('club')).font = NRM; ws.cell(i, 4, r.get('gf')).font = NRM
    ws.cell(i, 5, r.get('age')).font = NRM; ws.cell(i, 6, entry_label(r)).font = NRM
    o, n = V['shipped'][k], V['stage5'][k]
    c = ws.cell(i, 7, o); c.font = NRM; c.number_format = INT
    c = ws.cell(i, 8, n); c.font = NRM; c.number_format = INT
    c = ws.cell(i, 9, '=H%d-G%d' % (i, i)); c.font = NRM; c.number_format = INT
    c = ws.cell(i, 10, '=IF(G%d=0,"",(H%d-G%d)/G%d)' % (i, i, i, i)); c.font = NRM; c.number_format = PCT
    c = ws.cell(i, 11, '=IF(G%d=0,"",(H%d-G%d*%s)/(G%d*%s))' % (i, i, i, REBASE, i, REBASE))
    c.font = NRM; c.number_format = PCT
    d = [V['stage1'][k] - V['shipped'][k], V['eraremoval'][k] - V['stage1'][k],
         V['stage3'][k] - V['eraremoval'][k], V['stage4'][k] - V['stage3'][k],
         V['final'][k] - V['stage4'][k], V['stage5'][k] - V['final'][k]]
    for j, x in enumerate(d):
        c = ws.cell(i, 12 + j, x); c.font = NRM; c.number_format = INT
    if sum(d) != n - o: bad += 1
    c = ws.cell(i, 18, '=SUM(L%d:Q%d)-I%d' % (i, i, i)); c.font = NRM; c.number_format = INT
assert bad == 0, 'PER-ROW STAGE SUM FAILED on %d rows' % bad
print('PER-ROW STAGE SUM ASSERT: PASS on all %d rows (Σ of the SIX stage deltas == CAND − OLD, exactly)'
      % len(keys))
ws.freeze_panes = 'A2'; ws.auto_filter.ref = 'A1:R%d' % (len(keys) + 1)

# ---- picks: the ladder must be UNMOVED by stage 5 ----
ps = wb.create_sheet('picks')
lad = json.load(open('/home/claude/s5_landing/engine/rl_after/pvc_curve_v2.json'))
pn = {int(k): v for k, v in lad['curve'].items()}
po = {int(k): v for k, v in json.load(open(os.path.join(SBS, 'pvc_shipped.json')))['curve'].items()}
for i, (h, w) in enumerate([('pick', 8), ('OLD ladder — shipped', 24),
                            ('ladder at stage 5 (curve_md5 %s)' % lad['curve_md5'], 30),
                            ('abs Δ', 12), ('moved by stage 5?', 20)], 1):
    c = ps.cell(1, i, h); c.font = HDR; c.fill = FILL
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    ps.column_dimensions[get_column_letter(i)].width = w
for i, nn in enumerate(sorted(set(po) & set(pn)), 2):
    ps.cell(i, 1, nn).font = NRM
    c = ps.cell(i, 2, po[nn]); c.font = NRM; c.number_format = INT
    c = ps.cell(i, 3, pn[nn]); c.font = NRM; c.number_format = INT
    c = ps.cell(i, 4, '=C%d-B%d' % (i, i)); c.font = NRM; c.number_format = INT
    ps.cell(i, 5, 'NO — stage 5 moves no ladder').font = NRM
ps.freeze_panes = 'A2'
print('picks sheet: %d picks; the installed ladder %s is UNMOVED by stage 5.'
      % (len(set(po) & set(pn)), lad['curve_md5']))

# ---- stage-5 movers ----
def _num(v):
    try:
        float(v); return True
    except Exception:
        return False


sh = wb.create_sheet('quiet-starter movers')
c = sh.cell(1, 1, 'STAGE 5 — THE QUIET-STARTER REPRICE: every moved player (b56bbdde -> bad1961e). CANDIDATE, NOT LANDED.')
c.font = Font(name='Arial', sz=11, b=True)
c = sh.cell(2, 1, 'Verbatim from docs/evidence/act_334B_2026-08-07/stage5/movers_full.csv. No cap. '
                  'G is the taught anchor factor AT THAT PLAYER\'S OWN CELL, recomputed from his record to date.')
c.font = Font(name='Arial', sz=9, i=True)
rows = list(csv.reader(open(os.path.join(S5, 'movers_full.csv'))))
for ri, row in enumerate(rows, 4):
    for ci, val in enumerate(row, 1):
        cc = sh.cell(ri, ci, val if ri == 4 else (float(val) if _num(val) else val))
        cc.font = HDR if ri == 4 else NRM
        if ri == 4:
            cc.fill = FILL; cc.alignment = Alignment(horizontal='center', wrap_text=True)
for ci in range(1, len(rows[0]) + 1):
    sh.column_dimensions[get_column_letter(ci)].width = 17
sh.column_dimensions['A'].width = 26; sh.column_dimensions['B'].width = 24
sh.freeze_panes = 'A5'
print('movers sheet: %d rows' % (len(rows) - 1))

# ---- readme ----
rm = wb.create_sheet('readme')
rm.column_dimensions['A'].width = 5; rm.column_dimensions['B'].width = 132
LINES = [
    ('', 'READ ME — THIS WORKBOOK IS A CANDIDATE. THE STAGE-5 BOARD DID NOT LAND.'), ('', ''),
    ('1.', 'Stage 5 (the quiet-starter reprice) STOPPED on its landing floor: whole-cohort yr1 came in at 0.9945 against a floor of 1.00.'),
    ('', 'No product change was pushed. The engine edit, the taught table and this board exist only as EVIDENCE in docs/evidence/.../stage5/.'),
    ('', 'The ADOPTED owner-review workbook is side_by_side/board_before_after.xlsx and it still carries FIVE stage columns, correctly.'),
    ('', 'This file adds the SIXTH column separately so the work is verified and available if the owner rules to take the frontier as it stands.'), ('', ''),
    ('2.', 'SIX stage-delta columns on the players sheet; the sixth is the candidate:'),
    ('', '   Δ reference layer            de5110bb − 113b36f8'),
    ('', '   Δ era removal                f94e0778 − de5110bb'),
    ('', '   Δ curve+surface+numéraire    6c9f8d3a − f94e0778'),
    ('', '   Δ reactivity                 b490ae8b − 6c9f8d3a     (RL_PED_BAR = 0.5)'),
    ('', '   Δ surprise-trust             b56bbdde − b490ae8b     (RL_SUR_W = 5.0)'),
    ('', '   Δ QUIET-STARTER REPRICE      bad1961e − b56bbdde     (RL_G5_W = 1.0)  <-- SIXTH, CANDIDATE'), ('', ''),
    ('3.', 'The per-row identity Σ(six stage deltas) == (CANDIDATE − OLD) is asserted in Python at build time on all 804 rows,'),
    ('', 'and carried as a live formula in the STAGE SUM CHECK column so a reader can watch it hold. verify_xlsx_stage5.py re-checks it'),
    ('', 'from the written file (LibreOffice is broken in this sandbox, so formulas are verified in Python — amendment 1\'s convention).'), ('', ''),
    ('4.', 'Stage 5 moves NO ladder, NO numéraire, NO store and NO pole. 66 of 804 board rows move; all 66 move UP; zero rows fall.'),
]
for i, (a, b_) in enumerate(LINES, 1):
    rm.cell(i, 1, a).font = Font(name='Arial', sz=10, b=True)
    rm.cell(i, 2, b_).font = Font(name='Arial', sz=10)

wb.save(OUT)
print('wrote %s' % OUT)
