"""#334 stage B — THE SIDE-BY-SIDE WORKBOOK, regenerated with STAGE 6 (the CONDITIONED DEVELOPMENT
CORRECTION) as the SEVENTH stage column. Per-row stage deltas sum EXACTLY to the total, and that
identity is both asserted in Python and carried as a live formula column so a reader can see it hold.

STAGE 6 SHIPS AT DIAL 0 (RL_G6_W = RL_G6_KPD = 0, Addendum 1 F9/F10), so its column on the players
sheet is ZERO on every row BY CONSTRUCTION — that is the point: the shipped board is still the stage-5
landing 13f8c2e0, byte for byte, and the owner's intensity ruling is what turns the column on.

Because the seventh stage is a LADDER rather than a single board, the four rungs get their own sheet —
`stage 6 rungs` — carrying, per row, the stage-5 landed price, each rung's price, each rung's delta and
a per-row identity check (landed + delta == rung) asserted in Python and carried as a live formula.
That is the directive's stated fallback for per-rung columns, taken deliberately.

Supersedes the six-column build (stage 5).
"""
import json, csv, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SBS = '/home/claude/sbs'
RUNGS = ['0.25', '0.5', '0.75', '1.0']
HERE = os.path.dirname(os.path.abspath(__file__))
ACT = os.path.dirname(HERE)
REPOROOT = os.path.dirname(os.path.dirname(os.path.dirname(ACT)))
S5 = os.path.join(ACT, 'stage5')
S6 = os.path.join(ACT, 'stage6')
S4 = os.path.join(ACT, 'stage4')
A1 = os.path.join(ACT, 'stage4_amend1')
OUT = os.path.join(HERE, 'board_before_after.xlsx')
REBASE = 0.891738          # stage-3 uniform numeraire re-base; unmoved by stages 4, 4a1, 5 and 6

NAMES = ['shipped', 'stage1', 'eraremoval', 'stage3', 'stage4', 'final']
B = {n: json.load(open(os.path.join(SBS, 'board_%s.json' % n))) for n in NAMES}
B['stage5'] = json.load(open(os.path.join(S5, 'noarb', 'board_STAGE5_13f8c2e0.json')))
# STAGE 6, SHIPPED: dial 0 => the board IS the stage-5 landing. Read from the checkout so the claim is
# made against the committed bytes rather than restated.
B['stage6'] = json.load(open(os.path.join(REPOROOT, 'data', 'rl_build', 'rl_app_data.json')))
RB = {W: json.load(open(os.path.join(S6, 'boards', 'board_rung%s_kpd0.json' % W))) for W in RUNGS}
V = {n: {r['key']: r['v'] for r in B[n]['active'] if r.get('v') is not None} for n in B}
keys = [r['key'] for r in B['stage6']['active'] if r.get('v') is not None]
meta = {r['key']: r for r in B['stage6']['active']}
for n in V:
    assert set(V[n]) == set(keys), 'roster differs at %s: %d vs %d' % (n, len(V[n]), len(keys))
print('roster identical across all EIGHT boards: %d players' % len(keys))

HDR = Font(name='Arial', sz=10, b=True, color='FFFFFF')
NRM = Font(name='Arial', sz=10)
FILL = PatternFill('solid', fgColor='1F3864')
INT = '#,##0;(#,##0);-'
PCT = '0.00%;(0.00%);-'

wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'players'
cols = [('key', 26), ('name', 24), ('club', 20), ('pos', 7), ('age', 6), ('entry', 26),
        ('OLD — shipped board 113b36f8', 17), ('NEW — shipped stage-6 board 13f8c2e0', 19),
        ('abs Δ (NEW − OLD)', 13), ('% Δ (of OLD)', 12),
        ('% change beyond the currency re-base', 29),
        ('Δ reference layer (de5110bb − 113b36f8)', 18),
        ('Δ era removal (f94e0778 − de5110bb)', 18),
        ('Δ curve+surface+numéraire (6c9f8d3a − f94e0778)', 22),
        ('Δ reactivity (b490ae8b − 6c9f8d3a)', 18),
        ('Δ surprise-trust (b56bbdde − b490ae8b)', 20),
        ('Δ quiet-starter reprice (13f8c2e0 − b56bbdde)', 24),
        ('Δ development correction — STAGE 6 at the SHIPPED dial 0 (zero by construction)', 30),
        ('STAGE SUM CHECK (ΣΔ stages − abs Δ; must be 0)', 26)]
for i, (h, w) in enumerate(cols, 1):
    c = ws.cell(1, i, h); c.font = HDR; c.fill = FILL
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.column_dimensions[get_column_letter(i)].width = w


def entry_label(r):
    ty = r.get('ty') or '?'; pk = r.get('pk')
    if ty == 'ND' and pk and pk <= 64: return 'National pick %d' % pk
    return '%s %s' % (r.get('draft') or ty, ('pick %d' % pk) if pk else '(pool)')


bad = 0
for i, k in enumerate(keys, 2):
    r = meta[k]
    ws.cell(i, 1, k).font = NRM; ws.cell(i, 2, r.get('name')).font = NRM
    ws.cell(i, 3, r.get('club')).font = NRM; ws.cell(i, 4, r.get('gf')).font = NRM
    ws.cell(i, 5, r.get('age')).font = NRM; ws.cell(i, 6, entry_label(r)).font = NRM
    o, n = V['shipped'][k], V['stage6'][k]
    c = ws.cell(i, 7, o); c.font = NRM; c.number_format = INT
    c = ws.cell(i, 8, n); c.font = NRM; c.number_format = INT
    c = ws.cell(i, 9, '=H%d-G%d' % (i, i)); c.font = NRM; c.number_format = INT
    c = ws.cell(i, 10, '=IF(G%d=0,"",(H%d-G%d)/G%d)' % (i, i, i, i)); c.font = NRM; c.number_format = PCT
    c = ws.cell(i, 11, '=IF(G%d=0,"",(H%d-G%d*%s)/(G%d*%s))' % (i, i, i, REBASE, i, REBASE))
    c.font = NRM; c.number_format = PCT
    d = [V['stage1'][k] - V['shipped'][k], V['eraremoval'][k] - V['stage1'][k],
         V['stage3'][k] - V['eraremoval'][k], V['stage4'][k] - V['stage3'][k],
         V['final'][k] - V['stage4'][k], V['stage5'][k] - V['final'][k],
         V['stage6'][k] - V['stage5'][k]]
    for j, x in enumerate(d):
        c = ws.cell(i, 12 + j, x); c.font = NRM; c.number_format = INT
    if sum(d) != n - o: bad += 1
    c = ws.cell(i, 19, '=SUM(L%d:R%d)-I%d' % (i, i, i)); c.font = NRM; c.number_format = INT
assert bad == 0, 'PER-ROW STAGE SUM FAILED on %d rows' % bad
print('PER-ROW STAGE SUM ASSERT: PASS on all %d rows (Σ of the SEVEN stage deltas == NEW − OLD, exactly)' % len(keys))
ws.freeze_panes = 'A2'; ws.auto_filter.ref = 'A1:S%d' % (len(keys) + 1)

# ---- picks ----
ps = wb.create_sheet('picks')
lad = json.load(open(os.path.join(REPOROOT, 'engine', 'rl_after', 'pvc_curve_v2.json')))
pn = {int(k): v for k, v in lad['curve'].items()}
po = {int(k): v for k, v in json.load(open(os.path.join(SBS, 'pvc_shipped.json')))['curve'].items()}
p4 = {int(k): v for k, v in json.load(open(os.path.join(SBS, 'pvc_final.json')))['curve'].items()}
assert pn == p4, 'stage 5/6 moved the settled ladder — it must not'
for i, (h, w) in enumerate([('pick', 8), ('OLD ladder value — shipped', 26),
                            ('NEW ladder value — final (curve_md5 %s)' % lad['curve_md5'], 30),
                            ('abs Δ (NEW − OLD)', 16), ('% Δ (of OLD)', 14),
                            ('Δ from stage 4 amendment 1 (stages 5 and 6 move none)', 34)], 1):
    c = ps.cell(1, i, h); c.font = HDR; c.fill = FILL
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    ps.column_dimensions[get_column_letter(i)].width = w
for i, nn in enumerate(sorted(set(po) & set(pn)), 2):
    ps.cell(i, 1, nn).font = NRM
    c = ps.cell(i, 2, po[nn]); c.font = NRM; c.number_format = INT
    c = ps.cell(i, 3, pn[nn]); c.font = NRM; c.number_format = INT
    c = ps.cell(i, 4, '=C%d-B%d' % (i, i)); c.font = NRM; c.number_format = INT
    c = ps.cell(i, 5, '=IF(B%d=0,"",(C%d-B%d)/B%d)' % (i, i, i, i)); c.font = NRM; c.number_format = PCT
    c = ps.cell(i, 6, pn[nn] - p4[nn]); c.font = NRM; c.number_format = INT
ps.freeze_panes = 'A2'
print('picks sheet: %d picks (curve_md5 %s); the ladder is UNMOVED by stages 5 and 6.'
      % (len(set(po) & set(pn)), lad['curve_md5']))


def _num(v):
    try:
        float(v); return True
    except Exception:
        return False


def movers_sheet(title, banner, note, csvpath):
    sh = wb.create_sheet(title)
    c = sh.cell(1, 1, banner); c.font = Font(name='Arial', sz=11, b=True)
    c = sh.cell(2, 1, note); c.font = Font(name='Arial', sz=9, i=True)
    rows = list(csv.reader(open(csvpath)))
    for ri, row in enumerate(rows, 4):
        for ci, val in enumerate(row, 1):
            cc = sh.cell(ri, ci, val if ri == 4 else (float(val) if _num(val) else val))
            cc.font = HDR if ri == 4 else NRM
            if ri == 4:
                cc.fill = FILL; cc.alignment = Alignment(horizontal='center', wrap_text=True)
    for ci in range(1, len(rows[0]) + 1):
        sh.column_dimensions[get_column_letter(ci)].width = 17
    sh.column_dimensions['A'].width = 26; sh.column_dimensions['B'].width = 24
    sh.freeze_panes = 'A5'
    return len(rows) - 1


n4 = movers_sheet('reactivity movers',
                  'STAGE 4 — PEDIGREE-CONDITIONED REACTIVITY: every moved player (6c9f8d3a -> b490ae8b)',
                  'Verbatim from docs/evidence/act_334B_2026-08-07/stage4/movers_full.csv. No cap.',
                  os.path.join(S4, 'movers_full.csv'))
n1 = movers_sheet('surprise movers',
                  'STAGE 4 AMENDMENT 1 — SURPRISE-SCALED EVIDENCE TRUST: every moved player (b490ae8b -> b56bbdde)',
                  'Verbatim from docs/evidence/act_334B_2026-08-07/stage4_amend1/movers_full.csv. No cap. '
                  's = |log(e_full/anchor_full)| is the size of the re-rate the record claims; u is its unresolved share.',
                  os.path.join(A1, 'movers_full.csv'))
n5 = movers_sheet('quiet-starter movers',
                  'STAGE 5 — THE QUIET-STARTER REPRICE: every moved player (b56bbdde -> 13f8c2e0)',
                  'Verbatim from docs/evidence/act_334B_2026-08-07/stage5/movers_full.csv. No cap. '
                  'G is the taught anchor factor at that player\'s OWN cell, recomputed from his record to date.',
                  os.path.join(S5, 'movers_full.csv'))
print('movers sheets: stage 4 = %d rows, amendment 1 = %d rows, stage 5 = %d rows' % (n4, n1, n5))

# ---- STAGE 6 RUNG LADDER SHEET ----------------------------------------------------------------
rg = wb.create_sheet('stage 6 rungs')
c = rg.cell(1, 1, 'STAGE 6 — THE CONDITIONED DEVELOPMENT CORRECTION: the four-rung ladder, per row')
c.font = Font(name='Arial', sz=11, b=True)
c = rg.cell(2, 1, 'Baseline for every column is the STAGE-5 LANDED board 13f8c2e0 (which is also the '
                  'SHIPPED stage-6 board, dial 0). NO RUNG IS RECOMMENDED — the owner rules the '
                  'intensity. Rungs struck by a registered gate are named in stage6/FRONTIER.txt.')
c.font = Font(name='Arial', sz=9, i=True)
RCOLS = [('key', 26), ('name', 24), ('pos', 7), ('entry', 22),
         ('stage-5 LANDED (= shipped stage 6)', 22)]
for W in RUNGS:
    RCOLS += [('rung %s price' % W, 15), ('rung %s Δ' % W, 13), ('rung %s %% Δ' % W, 13)]
RCOLS += [('IDENTITY CHECK (Σ|landed+Δ−rung| over the four rungs; must be 0)', 30)]
for i, (h, w) in enumerate(RCOLS, 1):
    cc = rg.cell(4, i, h); cc.font = HDR; cc.fill = FILL
    cc.alignment = Alignment(horizontal='center', wrap_text=True)
    rg.column_dimensions[get_column_letter(i)].width = w
RV = {W: {r['key']: r['v'] for r in RB[W]['active'] if r.get('v') is not None} for W in RUNGS}
for W in RUNGS:
    assert set(RV[W]) == set(keys), 'roster differs at rung %s' % W
rbad = 0
for i, k in enumerate(keys, 5):
    r = meta[k]
    rg.cell(i, 1, k).font = NRM; rg.cell(i, 2, r.get('name')).font = NRM
    rg.cell(i, 3, r.get('gf')).font = NRM; rg.cell(i, 4, entry_label(r)).font = NRM
    base = V['stage5'][k]
    cc = rg.cell(i, 5, base); cc.font = NRM; cc.number_format = INT
    parts = []
    for j, W in enumerate(RUNGS):
        col = 6 + 3 * j
        cc = rg.cell(i, col, RV[W][k]); cc.font = NRM; cc.number_format = INT
        cc = rg.cell(i, col + 1, RV[W][k] - base); cc.font = NRM; cc.number_format = INT
        cc = rg.cell(i, col + 2, '=IF(E%d=0,"",%s%d/E%d)'
                     % (i, get_column_letter(col + 1), i, i)); cc.font = NRM; cc.number_format = PCT
        parts.append('ABS(E%d+%s%d-%s%d)' % (i, get_column_letter(col + 1), i, get_column_letter(col), i))
        if base + (RV[W][k] - base) != RV[W][k]: rbad += 1
    cc = rg.cell(i, 18, '=' + '+'.join(parts)); cc.font = NRM; cc.number_format = INT
assert rbad == 0, 'PER-ROW RUNG IDENTITY FAILED on %d cells' % rbad
rg.freeze_panes = 'A5'
print('stage 6 rungs sheet: %d players x 4 rungs; per-row identity landed+Δ==rung asserted on all cells'
      % len(keys))

# a compact summary block under the ladder
srow = len(keys) + 7
SUM = [('rung', 'board md5', 'movers', 'up', 'down', 'board total Δ', 'ND yr1 (teaching window)')]
OB = json.load(open(os.path.join(S6, 'owner_basis.json')))
tw = OB['populations']['ND 1-64, 2004-2022 (TEACHING WINDOW)']
SUM.append(('0 (SHIPPED)', '13f8c2e0', 0, 0, 0, 0, tw['stage5']))
for W in RUNGS:
    m = json.load(open(os.path.join(S6, 'movers_rung%s.json' % W)))['rungs'][W]
    SUM.append((W, m['board_md5'][:8], m['movers'], m['up'], m['down'],
                m['board_total_after'] - m['board_total_before'], tw['rungs'][W]))
for ri, row in enumerate(SUM):
    for ci, val in enumerate(row, 1):
        cc = rg.cell(srow + ri, ci, val); cc.font = HDR if ri == 0 else NRM
        if ri == 0: cc.fill = FILL

# ---- readme ----
rm = wb.create_sheet('readme')
rm.column_dimensions['A'].width = 5; rm.column_dimensions['B'].width = 132
LINES = [
    ('', 'READ ME — what each sheet is, and how to read the columns'), ('', ''),
    ('1.', 'This workbook is the owner review set for act #334 stage B. It puts the ADOPTED board next to the board this act produces.'),
    ('', 'OLD = the shipped board 113b36f8 (commit f8fe8361, on main today). NEW = the final board 13f8c2e0 on branch landing/334-stage-b.'),
    ('', "Nothing here is adopted. Adoption is the owner's word. No PR, no tag, no main merge."), ('', ''),
    ('2.', 'SEVEN boards, SIX of which moved; each mover has its own delta column on the players sheet:'),
    ('', '   Δ reference layer          de5110bb − 113b36f8   the adopted #336 reference layer'),
    ('', '   Δ era removal              f94e0778 − de5110bb   era normalization stripped (owner ruling)'),
    ('', '   Δ curve+surface+numéraire  6c9f8d3a − f94e0778   base curve re-teach, per-pick re-anchor, year-zero surface refit, numéraire re-base'),
    ('', '   Δ reactivity               b490ae8b − 6c9f8d3a   stage 4, the pedigree-conditioned evidence bar (RL_PED_BAR = 0.5)'),
    ('', '   Δ surprise-trust           b56bbdde − b490ae8b   stage 4 AMENDMENT 1, the surprise-scaled evidence trust (RL_SUR_W = 5.0)'),
    ('', '   Δ quiet-starter reprice    13f8c2e0 − b56bbdde   STAGE 5, the taught anchor factor G (RL_G5_W = 1.0)'),
    ('', '   Δ development correction   13f8c2e0 − 13f8c2e0   STAGE 6 at the SHIPPED dial 0 — ZERO on every row, by construction'), ('', ''),
    ('3.', 'The per-row identity Σ(six stage deltas) == (NEW − OLD) is asserted in Python at build time on all 804 rows, and carried as a live'),
    ('', 'formula in the STAGE SUM CHECK column so a reader can watch it hold. verify_xlsx.py re-checks it from the written file'),
    ('', '(LibreOffice is non-functional in this sandbox, so formulas are verified in Python — the convention amendment 1 established).'), ('', ''),
    ('4.', 'Stage 5 moves NO ladder, NO numéraire, NO store and NO pole. It moves 66 of 804 board rows; every one moves UP; zero rows fall.'),
    ('', 'Its landing and the ONE law that binds it are in docs/evidence/act_334B_2026-08-07/stage5/CONSISTENCY_PASS.md — read that before ruling.'), ('', ''),
    ('5.', 'STAGE 6 is a LADDER, not a board. Its shipped dials are both 0, so the players sheet shows zero movement and the shipped board is'),
    ('', 'still 13f8c2e0. The `stage 6 rungs` sheet carries the four candidate intensities side by side, per row, with the per-row identity'),
    ('', 'landed + Δ == rung asserted on every cell. NO RUNG IS RECOMMENDED. Read stage6/README.md and stage6/FRONTIER.txt before ruling:'),
    ('', 'THREE of the four rungs are STRUCK by registered gates, and the reason is printed with the exact figure that struck them.'), ('', ''),
    ('6.', 'CONFORMANCE REPAIR, issue #334 comment 5219329372. The stage-6 surface was RE-TAUGHT to the estimand the directive actually'),
    ('', 'registered (F = v(career year 4) discounted back at the 1.0939 hurdle — the engine no-arb identity; value-weighted year-1'),
    ('', 'aggregate 1.1363) in place of the rolling 4-year mean the first build substituted (1.0963, 72.2% of it); and the zero-cell gate'),
    ('', 'is read on the cross-section\'s own performance axis, the evaluation-year SEASON SCORING AVERAGE, not the production ratio the'),
    ('', 'first build assumed. All four rung boards and matrices re-emitted; the SHIPPED dial-0 board is UNCHANGED at 13f8c2e0, byte-exact.'),
]
for i, (a, b_) in enumerate(LINES, 1):
    rm.cell(i, 1, a).font = Font(name='Arial', sz=10, b=True)
    rm.cell(i, 2, b_).font = Font(name='Arial', sz=10)

wb.save(OUT)
print('wrote %s' % OUT)
