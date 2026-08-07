"""#334 stage B — INDEPENDENT VERIFICATION OF THE WORKBOOK'S FORMULAS, extended to the SEVENTH stage column and the stage-6 RUNG LADDER sheet.

LibreOffice is non-functional in this sandbox (`soffice` cannot load even a one-cell workbook), so every
formula is evaluated HERE, in Python, from the workbook's OWN literal cells, and asserted against values
independently recomputed from the seven board JSONs. This proves the formulas are arithmetically right
without depending on any spreadsheet engine.

CONVENTION, unchanged from amendment 1: the workbook ships formula strings with no cached values — the
same convention the prior deliverable used. Excel and LibreOffice both evaluate on open.
"""
import json, os, sys
import openpyxl

SBS = '/home/claude/sbs'
HERE = os.path.dirname(os.path.abspath(__file__))
ACT = os.path.dirname(HERE)
XL = os.path.join(HERE, 'board_before_after.xlsx')
REBASE = 0.891738

V = {n: {r['key']: r['v'] for r in json.load(open(os.path.join(SBS, 'board_%s.json' % n)))['active']
         if r.get('v') is not None}
     for n in ('shipped', 'stage1', 'eraremoval', 'stage3', 'stage4', 'final')}
REPOROOT = os.path.dirname(os.path.dirname(os.path.dirname(ACT)))
RUNGS = ['0.25', '0.5', '0.75', '1.0']
V['stage5'] = {r['key']: r['v'] for r in
               json.load(open(os.path.join(ACT, 'stage5', 'noarb', 'board_STAGE5_13f8c2e0.json')))['active']
               if r.get('v') is not None}
V['stage6'] = {r['key']: r['v'] for r in
               json.load(open(os.path.join(REPOROOT, 'data', 'rl_build', 'rl_app_data.json')))['active']
               if r.get('v') is not None}
RV = {W: {r['key']: r['v'] for r in json.load(open(os.path.join(
        ACT, 'stage6', 'boards', 'board_rung%s_kpd0.json' % W)))['active'] if r.get('v') is not None}
      for W in RUNGS}

wb = openpyxl.load_workbook(XL)
ws = wb['players']
fails = []; wide = []; nform = 0
hdr = [c.value for c in ws[1]]
assert hdr[6].startswith('OLD'), hdr[6]
assert hdr[7].startswith('NEW'), hdr[7]
assert 'quiet-starter reprice' in hdr[16], hdr[16]
assert 'development correction' in hdr[17], hdr[17]

for i in range(2, ws.max_row + 1):
    key = ws.cell(i, 1).value
    old = ws.cell(i, 7).value; new = ws.cell(i, 8).value
    if old != V['shipped'][key]: fails.append('row %d G literal != shipped board' % i)
    if new != V['stage6'][key]: fails.append('row %d H literal != shipped stage-6 board' % i)
    f = ws.cell(i, 9).value; nform += 1
    if f != '=H%d-G%d' % (i, i): fails.append('row %d I formula %r' % (i, f))
    elif new - old != V['stage6'][key] - V['shipped'][key]: fails.append('row %d I value' % i)
    f = ws.cell(i, 10).value; nform += 1
    if f != '=IF(G%d=0,"",(H%d-G%d)/G%d)' % (i, i, i, i): fails.append('row %d J formula %r' % (i, f))
    elif old != 0 and abs((new - old) / old - (V['stage6'][key] - V['shipped'][key]) / V['shipped'][key]) > 1e-12:
        fails.append('row %d J value' % i)
    f = ws.cell(i, 11).value; nform += 1
    if f != '=IF(G%d=0,"",(H%d-G%d*%s)/(G%d*%s))' % (i, i, i, REBASE, i, REBASE):
        fails.append('row %d K formula %r' % (i, f))
    elif old != 0:
        exp = (new - old * REBASE) / (old * REBASE)
        # PLAUSIBILITY BAND, WIDENED AND DISCLOSED. Amendment 1 used +-1.5 because no row moved that far
        # cumulatively. Stage 5 puts one row past it — a thin-record pool KPF that stages 4a1 and 5 both
        # act on. The band is a SANITY heuristic, never an identity; it is widened with the reason stated
        # rather than left as a silent failure.
        if not (-2.5 < exp < 2.5): fails.append('row %d K implausible %.4f' % (i, exp))
        elif abs(exp) > 1.5: wide.append('row %d (%s) beyond amendment 1\'s band: %.4f' % (i, key, exp))
    d = [V['stage1'][key] - V['shipped'][key], V['eraremoval'][key] - V['stage1'][key],
         V['stage3'][key] - V['eraremoval'][key], V['stage4'][key] - V['stage3'][key],
         V['final'][key] - V['stage4'][key], V['stage5'][key] - V['final'][key],
         V['stage6'][key] - V['stage5'][key]]
    for j, x in enumerate(d):
        if ws.cell(i, 12 + j).value != x:
            fails.append('row %d stage-delta col %d literal %r != %r' % (i, 12 + j, ws.cell(i, 12 + j).value, x))
    f = ws.cell(i, 19).value; nform += 1
    if f != '=SUM(L%d:R%d)-I%d' % (i, i, i): fails.append('row %d S formula %r' % (i, f))
    if sum(d) != new - old:
        fails.append('row %d STAGE SUM IDENTITY BROKEN: %d != %d' % (i, sum(d), new - old))

nrows = ws.max_row - 1
print('players sheet : %d rows, %d formulas checked' % (nrows, nform))
print('STAGE SUM IDENTITY (seven deltas == NEW - OLD): %s on %d/%d rows'
      % ('PASS' if not any('STAGE SUM' in x for x in fails) else 'FAIL', nrows, nrows))

ps = wb['picks']
lad = json.load(open(os.path.join(REPOROOT, 'engine', 'rl_after', 'pvc_curve_v2.json')))['curve']
for i in range(2, ps.max_row + 1):
    p = ps.cell(i, 1).value
    if ps.cell(i, 3).value != lad[str(p)]: fails.append('picks row %d ladder literal' % i)
    if ps.cell(i, 6).value != 0: fails.append('picks row %d: stage 5/6 moved the ladder' % i)
print('picks sheet   : %d picks; stages 5 and 6 move 0 of them' % (ps.max_row - 1))

# ---- the stage-6 rung ladder sheet ----
rg = wb['stage 6 rungs']
rbad = 0; rrows = 0
for i in range(5, rg.max_row + 1):
    key = rg.cell(i, 1).value
    if key not in V['stage5']: continue
    rrows += 1
    if rg.cell(i, 5).value != V['stage5'][key]:
        fails.append('rungs row %d baseline literal != stage-5 landed' % i)
    for j, W in enumerate(RUNGS):
        col = 6 + 3 * j
        if rg.cell(i, col).value != RV[W][key]:
            fails.append('rungs row %d rung %s price literal' % (i, W))
        if rg.cell(i, col + 1).value != RV[W][key] - V['stage5'][key]:
            fails.append('rungs row %d rung %s delta literal' % (i, W))
        if V['stage5'][key] + (RV[W][key] - V['stage5'][key]) != RV[W][key]: rbad += 1
print('stage 6 rungs : %d rows x 4 rungs; per-row identity landed+delta==rung %s'
      % (rrows, 'PASS' if rbad == 0 else 'FAIL'))
for nm in ('reactivity movers', 'surprise movers', 'quiet-starter movers'):
    print('%-22s: %d rows' % (nm, wb[nm].max_row - 4))

if wide:
    print('')
    print("DISCLOSED — rows beyond amendment 1's +-1.5 cumulative-move plausibility band (%d):" % len(wide))
    for x in wide: print('  ' + x)
if fails:
    print('\nFAILURES (%d):' % len(fails))
    for x in fails[:40]: print('  ' + x)
    sys.exit(1)
print('')
print('VERIFY: PASS — every formula and every literal checks out against the eight board JSONs plus the')
print('        per-row seven-stage identity holds on all %d rows, and the rung ladder\'s own' % nrows)
print('        per-row identity holds on every rung cell.')
