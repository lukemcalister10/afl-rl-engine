#!/usr/bin/env python3
"""
CONSOLIDATION STEP 1 — TASK 1: export the CURRENT source of truth (owner sanity-check).

READ-ONLY. Reads the authoritative store (rl_model_data.json = pre_stage0) and writes a
clean, human-readable xlsx: ONE ROW PER PLAYER, sorted alphabetically by player name,
SOURCE/input fields only (NO derived prices).

Re-runnable: `python3 evidence/consolidation_map/export_source_of_truth.py`

Ground truth asserted at run time:
  git HEAD  389ac39  (tag baked-v2.4-2026-07-04)
  store md5 644d1254  (rl_model_data.json, pre_stage0, authoritative)

DATA_SOURCE for every column = engine/rl_after/rl_model_data.json
"""
import json, os, sys, hashlib, subprocess, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
STORE = os.path.join(ROOT, "engine/rl_after/rl_model_data.json")
OUT   = os.path.join(ROOT, "evidence/consolidation_map/source_of_truth_current.xlsx")

EXPECT_SHA = "389ac39"
EXPECT_MD5 = "644d1254"

def md5_8(path):
    return hashlib.md5(open(path, "rb").read()).hexdigest()[:8]

def assert_ground_truth():
    try:
        sha = subprocess.check_output(["git", "-C", ROOT, "rev-parse", "--short", "HEAD"]).decode().strip()
    except Exception as e:
        sha = f"(git error: {e})"
    md5 = md5_8(STORE)
    print(f"[ground truth] git HEAD short = {sha}  (expect {EXPECT_SHA})")
    print(f"[ground truth] store md5(8)   = {md5}  (expect {EXPECT_MD5})")
    if not sha.startswith(EXPECT_SHA):
        print("!! WARNING: HEAD does not match baked head 389ac39", file=sys.stderr)
    if md5 != EXPECT_MD5:
        print("!! WARNING: store md5 does not match authoritative 644d1254", file=sys.stderr)

def recent_avg(rec):
    """Most-recent-season avg = the scoring entry with the highest year."""
    sc = rec.get("scoring") or []
    if not sc:
        return (None, None, None)
    latest = max(sc, key=lambda s: s.get("year", -1))
    return (latest.get("year"), latest.get("avg"), latest.get("games"))

def present_position(rec):
    """present position = _pos_now override if set, else drafted pos."""
    return rec.get("_pos_now") or rec.get("pos")

def surname(name):
    base = name.split("(")[0].strip()           # strip '(IRE credit)' etc.
    parts = base.split()
    return parts[-1] if parts else base

# ---- source column spec: label -> extractor. SOURCE/input fields only. ----
COLUMNS = [
    ("name",                 lambda r: r.get("player")),
    ("key_id",               lambda r: r.get("key")),
    ("DOB",                  lambda r: r.get("_bd")),
    ("born_year",            lambda r: r.get("_by")),
    ("draft_year",           lambda r: r.get("year")),
    ("pick",                 lambda r: r.get("pick")),
    ("pickless_flag",        lambda r: r.get("_pickless")),
    ("draft_type",           lambda r: r.get("type")),
    ("draft_category__cat",  lambda r: r.get("_cat")),
    ("draft_channel__draft", lambda r: r.get("_draft")),
    ("club",                 lambda r: r.get("_club")),
    ("present_position",     present_position),
    ("drafted_position__pos",lambda r: r.get("pos")),
    ("pos_now_override",     lambda r: r.get("_pos_now")),
    ("retired_flag",         lambda r: r.get("_retired")),
    ("career_games",         lambda r: r.get("games")),
    ("recent_season_year",   lambda r: recent_avg(r)[0]),
    ("recent_season_avg",    lambda r: recent_avg(r)[1]),
    ("recent_season_games",  lambda r: recent_avg(r)[2]),
    ("has_2026_season",      lambda r: r.get("_has26")),
    ("force_active_flag",    lambda r: r.get("_force_active")),
    ("last_listed_year",     lambda r: r.get("_last_listed")),
    ("raw_multipos__fut",    lambda r: json.dumps(r.get("_fut"), ensure_ascii=False) if r.get("_fut") is not None else None),
    ("phantom_flag",         lambda r: r.get("_phantom")),
    ("double_count_flag",    lambda r: r.get("_double_count")),
    ("pvc_exclude_flag",     lambda r: r.get("_pvc_exclude")),
]

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF")
COLL_FILL = PatternFill("solid", fgColor="7F6000")

def style_header(ws, ncols, fill=HDR_FILL):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

def main():
    print("=" * 60)
    print("TASK 1 — export source of truth (READ-ONLY)")
    print("=" * 60)
    assert_ground_truth()

    data = json.load(open(STORE, encoding="utf-8"))
    print(f"[store] path      = {STORE}")
    print(f"[store] byte-size = {os.path.getsize(STORE):,}")
    print(f"[store] records   = {len(data)}")

    rows = sorted(data, key=lambda r: (r.get("player") or "").lower())

    wb = Workbook()
    ws = wb.active
    ws.title = "SOURCE_current"
    ws.append([c[0] for c in COLUMNS])
    for r in rows:
        ws.append([fn(r) for _, fn in COLUMNS])
    style_header(ws, len(COLUMNS))
    # reasonable widths
    for i, (label, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(10, min(28, len(label) + 4))

    # ---- COLLISIONS tab: players sharing a surname, keyed by ID ----
    groups = collections.defaultdict(list)
    for r in data:
        groups[surname(r.get("player") or "").lower()].append(r)
    coll = {s: rs for s, rs in groups.items() if len(rs) > 1}

    ws2 = wb.create_sheet("COLLISIONS")
    ws2.append(["surname", "name", "key_id", "pick", "draft_year",
                "draft_type", "club", "drafted_position__pos", "present_position"])
    n_players = 0
    for s in sorted(coll):
        members = sorted(coll[s], key=lambda r: (r.get("year") or 0))
        for r in members:
            n_players += 1
            ws2.append([s, r.get("player"), r.get("key"), r.get("pick"), r.get("year"),
                        r.get("type"), r.get("_club"), r.get("pos"), present_position(r)])
    style_header(ws2, 9, COLL_FILL)
    for i, w in enumerate([16, 24, 26, 8, 11, 11, 20, 18, 16], start=1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w

    wb.save(OUT)
    print(f"\n[export] wrote {OUT}")
    print(f"[export] SOURCE_current rows = {len(rows)}  (one per player)")
    print(f"[export] COLLISIONS: {len(coll)} shared surnames, {n_players} players")
    print(f"[export] DATA_SOURCE (all figures) = {os.path.relpath(STORE, ROOT)}")

if __name__ == "__main__":
    main()
