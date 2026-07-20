#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
UI v1.2 — THE INPUTS INGEST (items 178(2)/(3)/(4), 180, 181(2)).  Directive:
docs/DIRECTIVE_UI_v1_2_club_valuation_2026-07-16.md.

The no-LLM pipeline's deterministic half: read the owner's authored input files under docs/inputs/
STRICTLY READ-ONLY, VALIDATE-OR-HALT (halt, never guess/warn-and-continue), price every held draft
pick off the SHIPPED engine's canonical pick curve, aggregate a per-club valuation, and emit the
committed view bundle ui/data/club_valuation.js (window.__CLUB_VALUATION__ = {...}).

FENCE: ui/-only + read-only reads of docs/inputs/ and the shipped board bundle + the engine curve
file (for the stamp cross-check only).  NO value is recomputed here — a pick's price is a MEAN of the
engine's own curve; a player's value is the stamped board `v`, only summed/greedily selected.  Values
from the pick workbook (Raw Value / Value (counted) / Pick Values tab) are NEVER ingested.

Run:  python3 ui/tools/ingest_inputs.py         (exit 0 = clean bundle written; exit 2 = HALT)
"""
import csv, json, os, sys, collections, hashlib, datetime

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # repo root
# Production paths are the defaults; each may be redirected via an env var so a FIXTURE run (a temp board
# bundle / temp contract / temp engine-curve dir / temp boot manifest / temp out) exercises the SAME
# fail-closed resolver and asserts without touching production data. Overrides change WHERE we read/write,
# never WHAT we assert — every ring-fence + curve-provenance HALT below stays live regardless of the paths.
INPUTS = os.environ.get("RL_UI_INPUTS", os.path.join(ROOT, "docs", "inputs"))
UI_DATA = os.path.join(ROOT, "ui", "data")
BOARD_BUNDLE = os.environ.get("RL_UI_BOARD_BUNDLE", os.path.join(UI_DATA, "board_view_working.js"))
# Engine curve DIR (not a single hardcoded file): the release-active curve is resolved from the contract.
ENGINE_DIR = os.environ.get("RL_UI_ENGINE_DIR", os.path.join(ROOT, "engine", "rl_after"))
# The explicit, fail-closed release-metadata contract that declares the release-active pick curve.
CURVE_CONTRACT = os.environ.get("RL_UI_CURVE_CONTRACT", os.path.join(ROOT, "ui", "release_pick_curve.json"))
# The accepted release manifest — read STRICTLY READ-ONLY for the store + release_version cross-check.
BOOT = os.environ.get("RL_UI_BOOT", os.path.join(ROOT, "data", "expected_boot.json"))
OUT = os.environ.get("RL_UI_OUT", os.path.join(UI_DATA, "club_valuation.js"))

EXPECTED_BOARD = "06d8af60"      # ui/app/config.js EXPECTED_BOARD — v2.11-rc1 balanced board of record
PICK_FUTURE_DISCOUNT = 0.10      # R104.5 balanced posture — the ONLY posture in this build
BASE_YEAR = 2026

# Known release pick-curve pathways -> the engine curve FILENAME each pathway loads. The contract's
# adopted_pathway MUST be one of these; the resolver cross-checks the contract's declared path AND the
# curve file's OWN self-declared gate token against this registry. An unknown pathway HALTs.
KNOWN_PATHWAYS = {
    "RL_PVC2": "pvc_curve_v2.json",       # v2.9+ composed pathway (the v2.11 adopted curve)
    "RL_PVCADOPT": "pvc_curve_L1b.json",  # prior v2.9 L1b adopt curve (superseded by RL_PVC2)
}
# Best-23 positional structure (item 178(3)); posCode vocab is the board's CURRENT posCode.
SLOTS = [("KEY_DEF", 2), ("GEN_DEF", 4), ("MID", 5), ("GEN_FWD", 4), ("KEY_FWD", 2), ("RUC", 1)]
BENCH = 5
FREE_AGENTS = "Free Agents"

verdicts = []   # [(check, ok, detail)]
notes = []


def check(name, ok, detail=""):
    verdicts.append((name, bool(ok), detail))
    return bool(ok)


class HaltError(Exception):
    """Raised by halt(); caught in main() -> renders the verdict table, writes a HALTED overlay bundle
    (so the UI refuses to render) and exits 2. Making halt() RAISE rather than sys.exit lets the
    curve-provenance resolver + guards be unit-tested for fail-closed behaviour (import + assert-raises)."""

    def __init__(self, reason):
        self.reason = reason
        super().__init__(reason)


def halt(reason):
    """Fail closed. Nothing is ever guessed; the overlay refuses to render on an ambiguous ingest."""
    raise HaltError(reason)


def _emit_halt(reason):
    """Render the verdict table + the halt reason and write the HALTED bundle (overlay refuses). Called
    once, from main(), when any guard raises HaltError."""
    _print_verdicts()
    print("\n■ HALT — %s" % reason)
    print("  The club-valuation overlay refuses to render on this ingest.  Nothing is guessed.")
    payload = {
        "stamp": {"expectedBoard": EXPECTED_BOARD, "generated": _now()},
        "halt": {"reason": reason, "verdicts": [{"check": c, "ok": o, "detail": d} for c, o, d in verdicts]},
        "verdicts": [{"check": c, "ok": o, "detail": d} for c, o, d in verdicts],
        "clubs": [], "picksByTeam": {}, "notes": notes,
    }
    _write(payload)


def _now():
    # deterministic-enough provenance stamp; no bearing on any value
    return datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")


def readcsv(path):
    for enc in ("utf-8", "cp1252", "latin-1"):
        try:
            with open(path, encoding=enc) as f:
                return list(csv.reader(f)), enc
        except UnicodeDecodeError:
            continue
    halt("could not decode %s as utf-8/cp1252/latin-1" % os.path.basename(path))


def nkey(s):
    """Normalised name join key: collapse whitespace + casefold.  Resolves the 5 case-only surname
       variants (MacDonald/Macdonald, DeMattia/Demattia, …) between the board and the owner's CSVs
       without guessing; still keeps the two Max Kings distinct ('max king' != 'maxwell king')."""
    return " ".join(str(s).strip().split()).casefold()


# ---------------------------------------------------------------------------- load the shipped board
def load_board():
    if not os.path.exists(BOARD_BUNDLE):
        halt("board bundle missing: %s" % BOARD_BUNDLE)
    s = open(BOARD_BUNDLE, encoding="utf-8").read()
    try:
        obj = json.loads(s[s.index("{"): s.rindex("}") + 1])
    except Exception as e:
        halt("board bundle is not parseable JSON: %s" % e)
    stamp = obj.get("stamp", {})
    board = str(stamp.get("board", ""))
    check("board id ring-fence", board[:8] == EXPECTED_BOARD,
          "bundle board %s vs expected %s" % (board[:8], EXPECTED_BOARD))
    if board[:8] != EXPECTED_BOARD:
        halt("board id mismatch — bundle %s != EXPECTED_BOARD %s (regenerate board_view or re-pin)"
             % (board[:8], EXPECTED_BOARD))
    pvc = {int(k): v for k, v in obj.get("pvc", {}).items()}
    return obj, stamp, pvc


def _md5_file(path):
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 16), b""):
            h.update(chunk)
    return h.hexdigest()


def _read_json(path, what):
    if not os.path.exists(path):
        halt("%s missing: %s (HALT-AND-ASK on provenance)" % (what, path))
    try:
        return json.load(open(path, encoding="utf-8"))
    except Exception as e:
        halt("%s is not parseable JSON: %s (%s)" % (what, path, e))


def _gate_token(curve_doc):
    """The first whitespace token of a curve file's self-declared 'gate' string is its pathway id
    (e.g. 'RL_PVC2 (parallel of RL_PVCADOPT) ...' -> 'RL_PVC2'). Absent gate -> None (fails closed later)."""
    g = str(curve_doc.get("gate", "")).strip()
    return g.split()[0] if g else None


# --------------------------------------------- release-active pick curve: deterministic + fail-closed (S5)
def load_curve_contract():
    """Read the EXPLICIT release-metadata contract (ui/release_pick_curve.json) and cross-check it against
    the ACCEPTED release manifest (data/expected_boot.json, read-only): store + release_version must agree,
    pin1 must be 3000, adopted_pathway must be KNOWN. Missing / unknown / conflicting -> HALT. This is the
    single deterministic source of the release-active pathway: the config manifest pins RL_PVCADOPT but does
    NOT carry the RL_PVC2 default-ON kill-switch, so the pathway cannot be read from the config alone."""
    c = _read_json(CURVE_CONTRACT, "release pick-curve contract")
    need = ("release_version", "as_of_round", "adopted_pathway", "pick_curve_path",
            "pick_curve_file_md5", "pick_curve_curve_md5", "store_md5", "numeraire_pin1")
    miss = [k for k in need if k not in c]
    check("release pick-curve contract has all required fields", not miss, "missing: %s" % (miss or "none"))
    if miss:
        halt("release pick-curve contract is incomplete (missing %s) — cannot resolve the release-active "
             "curve (HALT-AND-ASK)" % miss)
    pathway = c["adopted_pathway"]
    check("contract adopted_pathway is a known pathway", pathway in KNOWN_PATHWAYS,
          "%s (known: %s)" % (pathway, sorted(KNOWN_PATHWAYS)))
    if pathway not in KNOWN_PATHWAYS:
        halt("UNKNOWN curve-selection: adopted_pathway '%s' is not a known pathway %s — refusing to guess "
             "(HALT-AND-ASK)" % (pathway, sorted(KNOWN_PATHWAYS)))
    boot = _read_json(BOOT, "release manifest (expected_boot)")
    store_ok = str(c["store_md5"]) == str(boot.get("store"))
    check("contract store_md5 == release manifest store", store_ok,
          "contract %s vs manifest %s" % (str(c["store_md5"])[:8], str(boot.get("store"))[:8]))
    if not store_ok:
        halt("CONFLICTING curve-selection: contract store_md5 %s != release store %s — this pick-curve "
             "contract is not for this release (HALT-AND-ASK)"
             % (str(c["store_md5"])[:8], str(boot.get("store"))[:8]))
    rv_ok = str(c["release_version"]) == str(boot.get("release_version"))
    check("contract release_version == release manifest release_version", rv_ok,
          "contract %s vs manifest %s" % (c["release_version"], boot.get("release_version")))
    if not rv_ok:
        halt("CONFLICTING curve-selection: contract release_version %s != release %s (HALT-AND-ASK)"
             % (c["release_version"], boot.get("release_version")))
    # as_of_round must be present on the contract AND agree with the release manifest — a contract for a
    # different round is a stale/wrong pick-curve selection and must fail closed (not silently priced).
    boot_round = boot.get("as_of_round")
    ar_ok = boot_round is not None and str(c["as_of_round"]) == str(boot_round)
    check("contract as_of_round == release manifest as_of_round", ar_ok,
          "contract %s vs manifest %s" % (c["as_of_round"], boot_round))
    if not ar_ok:
        halt("CONFLICTING curve-selection: contract as_of_round %s != release as_of_round %s — the pick-curve "
             "contract is not for this release round (HALT-AND-ASK)" % (c["as_of_round"], boot_round))
    pin_ok = int(c["numeraire_pin1"]) == 3000
    check("contract numeraire pin1 == 3000", pin_ok, "pin1 = %s" % c["numeraire_pin1"])
    if not pin_ok:
        halt("contract numeraire pin1 != 3000 — numeraire drift; refusing to resolve a drifted ruler")
    return c


def resolve_release_curve(contract):
    """Resolve the release-active engine pick curve DETERMINISTICALLY from the contract, cross-checking the
    curve file's OWN self-declared identity: the filename the pathway must load, full-file md5, curve_md5,
    gate token, and store binding. Any drift/conflict -> HALT. Returns {'curve': {pick:int}, 'gate','path',...}."""
    pathway = contract["adopted_pathway"]
    want_name = KNOWN_PATHWAYS[pathway]
    got_name = os.path.basename(str(contract["pick_curve_path"]))
    name_ok = got_name == want_name
    check("contract pick_curve_path matches the adopted pathway", name_ok,
          "%s expects %s, contract points at %s" % (pathway, want_name, got_name))
    if not name_ok:
        halt("CONFLICTING curve-selection: adopted_pathway %s must load %s but the contract points at %s "
             "(e.g. L1b supplied while RL_PVC2 is active) — HALT-AND-ASK" % (pathway, want_name, got_name))
    path = os.path.join(ENGINE_DIR, want_name)
    doc = _read_json(path, "release-active engine curve %s" % want_name)
    file_md5 = _md5_file(path)
    md5_ok = file_md5 == str(contract["pick_curve_file_md5"])
    check("release-active curve file md5 == contract", md5_ok,
          "file %s vs contract %s" % (file_md5[:8], str(contract["pick_curve_file_md5"])[:8]))
    if not md5_ok:
        halt("CURVE DRIFT: %s md5 %s != contract %s — the engine curve file changed under the contract "
             "(HALT-AND-ASK)" % (want_name, file_md5[:8], str(contract["pick_curve_file_md5"])[:8]))
    cmd5_ok = str(doc.get("curve_md5")) == str(contract["pick_curve_curve_md5"])
    check("release-active curve curve_md5 == contract", cmd5_ok,
          "curve %s vs contract %s" % (doc.get("curve_md5"), contract["pick_curve_curve_md5"]))
    if not cmd5_ok:
        halt("CURVE DRIFT: %s curve_md5 %s != contract %s (HALT-AND-ASK)"
             % (want_name, doc.get("curve_md5"), contract["pick_curve_curve_md5"]))
    gate = _gate_token(doc)
    gate_ok = gate == pathway
    check("engine curve self-declares gate == adopted pathway", gate_ok,
          "curve gate '%s' vs pathway '%s'" % (gate, pathway))
    if not gate_ok:
        halt("CONFLICTING curve-selection: %s self-declares gate '%s' != adopted pathway '%s' — refusing "
             "to price on a curve from the wrong pathway (HALT-AND-ASK)" % (want_name, gate, pathway))
    # If the curve self-declares a store binding it must equal the release store (v2 carries stamp.store_md5;
    # L1b does not — an ABSENT binding is permitted, a PRESENT-and-WRONG one HALTs).
    curve_store = str(((doc.get("stamp") or {}).get("store_md5")) or "")
    if curve_store:
        cs_ok = curve_store[:8] == str(contract["store_md5"])[:8]
        check("engine curve store binding == release store", cs_ok,
              "curve %s vs release %s" % (curve_store[:8], str(contract["store_md5"])[:8]))
        if not cs_ok:
            halt("CONFLICTING curve-selection: %s binds store %s != release store %s (HALT-AND-ASK)"
                 % (want_name, curve_store[:8], str(contract["store_md5"])[:8]))
    if int(doc.get("pin", 0)) != 3000:
        halt("release-active curve %s pin != 3000 — numeraire drift (HALT-AND-ASK)" % want_name)
    eng = {int(k): int(v) for k, v in doc["curve"].items()}
    return {"curve": eng, "gate": pathway, "path": "engine/rl_after/" + want_name,
            "file_md5": file_md5, "curve_md5": str(doc.get("curve_md5")), "store_md5": str(contract["store_md5"])}


# ---------------------------------------- board PVC == release-active curve (S5 stamp-assert, corrected)
def assert_pvc(pvc, resolved):
    """Cross-check the INSTALLED board's PVC against the release-active curve resolved above: full shared-pick
    byte equality + pin1 + monotone. A board built on any OTHER pathway (e.g. an L1b / RL_PVC2=0 board while
    the contract adopts RL_PVC2) fails this equality and HALTs. Same S5 doctrine, on the CORRECT curve."""
    if not pvc:
        halt("no PVC in the board bundle — cannot locate the canonical pick curve (HALT-AND-ASK)")
    eng = resolved["curve"]
    shared = [k for k in pvc if k in eng]
    bytematch = bool(shared) and all(pvc[k] == eng[k] for k in shared)
    check("board PVC == release-active %s (%s, over %d shared picks)"
          % (resolved["path"], resolved["gate"], len(shared)), bytematch)
    if not bytematch:
        halt("STALE-CURVE GUARD: the board's PVC does not byte-match the release-active curve %s "
             "(%s, curve_md5 %s) — provenance ambiguous (the S5 failure). HALT-AND-ASK."
             % (resolved["path"], resolved["gate"], resolved["curve_md5"]))
    check("PVC numeraire anchor pick1 == 3000", pvc.get(1) == 3000, "pick1 = %s" % pvc.get(1))
    if pvc.get(1) != 3000:
        halt("PVC pick1 != 3000 — numeraire drift; refusing to price picks against a drifted ruler")
    # monotone non-increasing (the curve is a ruler)
    ks = sorted(pvc)
    mono = all(pvc[ks[i]] >= pvc[ks[i + 1]] for i in range(len(ks) - 1))
    check("PVC monotone non-increasing", mono)
    if not mono:
        halt("PVC is not monotone non-increasing — not a valid pick ruler")
    return pvc


# ----------------------------------------------------------------------------------- price one pick
def price_pick(pvc, lo, hi, year):
    vals = [pvc[p] for p in range(lo, hi + 1)]
    mean = sum(vals) / len(vals)
    if year == 2027:
        mean *= (1.0 - PICK_FUTURE_DISCOUNT)
    return mean


# ----------------------------------------------------------------------------------------- the picks
def load_picks(pvc, affl_teams):
    try:
        import openpyxl
    except ImportError:
        halt("openpyxl not available — cannot read AFFL_Pick_Locations.xlsx")
    path = os.path.join(INPUTS, "AFFL_Pick_Locations.xlsx")
    if not os.path.exists(path):
        halt("pick workbook missing: %s" % path)
    import warnings
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(path, data_only=True)

    # --- Ladder: the 2027 multiplier cell (read + reconcile against R104.5's governing 0.10) ---
    mult = None
    if "Ladder" in wb.sheetnames:
        for row in wb["Ladder"].iter_rows(values_only=True):
            if row and row[0] and str(row[0]).startswith("2027 value multiplier"):
                mult = row[1]
    governing = round(1.0 - PICK_FUTURE_DISCOUNT, 6)
    agree = mult is not None and abs(float(mult) - governing) < 1e-6
    check("ladder 2027 multiplier reconciles to R104.5 balanced (1-0.10=0.90)", agree,
          "sheet cell = %s · governing = %s" % (mult, governing))
    if not agree:
        halt("2027 MULTIPLIER DISAGREEMENT: the sheet says %s but R104.5 balanced governs %s — "
             "reconcile the sheet or the ruling (HALT-AND-FLAG, never silently pick)" % (mult, governing))

    # --- Picks ledger ---
    if "Picks" not in wb.sheetnames:
        halt("pick workbook has no 'Picks' sheet")
    picks = []
    for i, row in enumerate(wb["Picks"].iter_rows(values_only=True)):
        if i < 2 or row[0] is None:   # rows 0-1 are title+header
            continue
        pid, yr, rnd, orig, owner, lo, hi = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
        picks.append(dict(id=pid, year=int(yr), rnd=rnd, orig=str(orig).strip(),
                          owner=str(owner).strip(), lo=int(lo), hi=int(hi)))

    check("pick ledger row count == 160", len(picks) == 160, "%d rows" % len(picks))

    # bands 1..80, low <= high
    band_ok = all(1 <= p["lo"] <= 80 and 1 <= p["hi"] <= 80 and p["lo"] <= p["hi"] for p in picks)
    bad = [p["id"] for p in picks if not (1 <= p["lo"] <= 80 and 1 <= p["hi"] <= 80 and p["lo"] <= p["hi"])]
    check("all bands within 1-80 and low <= high", band_ok, "offenders: %s" % (bad or "none"))
    if not band_ok:
        halt("band violation on pick ids %s (out of 1-80 or low>high)" % bad)

    # draft year <= one year ahead
    yrs = sorted({p["year"] for p in picks})
    yr_ok = all(BASE_YEAR <= y <= BASE_YEAR + 1 for y in yrs)
    check("both drafts <= one year ahead (%d/%d)" % (BASE_YEAR, BASE_YEAR + 1), yr_ok, "years=%s" % yrs)
    if not yr_ok:
        halt("draft year out of [%d, %d]: %s" % (BASE_YEAR, BASE_YEAR + 1, yrs))

    # every Owner + Origin joins to exactly one AFFL club (Free Agents permitted)
    club_set = set(affl_teams) | {FREE_AGENTS}

    def join_club(short):
        if short in club_set:
            return [short]
        cand = [t for t in affl_teams if t == short or t.startswith(short + " ")]
        return cand

    omap = {}
    for field in ("owner", "orig"):
        for p in picks:
            v = p[field]
            c = join_club(v)
            if len(c) != 1:
                check("%s '%s' joins to exactly one AFFL club" % (field, v), False,
                      "candidates: %s" % c)
                halt("%s '%s' does not join to a unique AFFL club (candidates %s)" % (field, v, c))
            omap[v] = c[0]
    check("every pick Owner + Origin joins to exactly one AFFL club (+Free Agents)", True,
          "%d distinct owners+origins" % len(omap))

    # price + attach the joined team
    for p in picks:
        p["team"] = omap[p["owner"]]
        p["origin_team"] = omap[p["orig"]]
        p["value"] = round(price_pick(pvc, p["lo"], p["hi"], p["year"]))
        p["band"] = "#%d-%d" % (p["lo"], p["hi"]) if p["lo"] != p["hi"] else "#%d" % p["lo"]

    # pick-count conservation (the Dashboard convention): ledger == Sum per-owner counts == 160
    per_owner = collections.Counter(p["team"] for p in picks)
    conserved = sum(per_owner.values()) == len(picks) == 160
    check("pick-count conservation (Sum per-club == ledger == 160)", conserved,
          "sum=%d ledger=%d" % (sum(per_owner.values()), len(picks)))
    if not conserved:
        halt("pick-count conservation failed")
    check("all 160 picks priced off the canonical curve", all("value" in p for p in picks),
          "priced=%d" % sum(1 for p in picks if "value" in p))
    return picks


# ---------------------------------------------------------------- the CSV validations (name -> id)
def validate_csvs(board_by_nkey):
    loc, lenc = readcsv(os.path.join(INPUTS, "AFFL_Player_Locations.csv"))
    pos, penc = readcsv(os.path.join(INPUTS, "AFFL_Future_Positioning.csv"))
    loc_rows = [r for r in loc[1:] if r and r[0]]
    pos_rows = [r for r in pos[1:] if r and r[0]]

    # normalised-key uniqueness within each source (zero ambiguity)
    loc_by = collections.defaultdict(list)
    for r in loc_rows:
        loc_by[nkey(r[0])].append(r)
    pos_by = collections.defaultdict(list)
    for r in pos_rows:
        pos_by[nkey(r[1])].append(r)   # positioning col1 = player_name, col0 = stable_player_id
    loc_dups = [k for k, v in loc_by.items() if len(v) > 1]
    pos_dups = [k for k, v in pos_by.items() if len(v) > 1]
    check("player names unambiguous in Player_Locations.csv (normalised)", not loc_dups,
          "dups: %s" % (loc_dups or "none"))
    check("player names unambiguous in Future_Positioning.csv (normalised)", not pos_dups,
          "dups: %s" % (pos_dups or "none"))
    if loc_dups or pos_dups:
        halt("ambiguous player name(s) in the input CSVs: %s" % (loc_dups + pos_dups))

    # every location name joins to exactly one stable id via positioning, and to exactly one board row
    unmatched_pos = [r[0] for r in loc_rows if nkey(r[0]) not in pos_by]
    unmatched_board = [r[0] for r in loc_rows if nkey(r[0]) not in board_by_nkey]
    check("every Player_Locations name -> a unique stable_player_id (positioning join)",
          not unmatched_pos, "unmatched: %s" % (unmatched_pos or "none"))
    check("every Player_Locations name -> a unique board player (name join)",
          not unmatched_board, "unmatched: %s" % (unmatched_board or "none"))
    if unmatched_pos or unmatched_board:
        halt("player name(s) fail the id/board join: %s" % (unmatched_pos + unmatched_board))

    # the two Max Kings asserted distinct by name
    mk, mxk = nkey("Max King"), nkey("Maxwell King")
    distinct = (mk != mxk and mk in pos_by and mxk in pos_by
                and pos_by[mk][0][0] != pos_by[mxk][0][0])
    check("the two Max Kings are distinct by name (assert)", distinct,
          "Max King=%s · Maxwell King=%s" % (
              pos_by.get(mk, [["?"]])[0][0], pos_by.get(mxk, [["?"]])[0][0]))
    if not distinct:
        halt("the two Max Kings failed the distinctness assertion")

    # 5 case-only surname variants: reconciled by the normalised key, recorded (not a halt)
    exact_board_names = {rows[0]["name"].strip() for rows in board_by_nkey.values()}
    variants = sorted({r[0] for r in loc_rows
                       if r[0].strip() not in exact_board_names and nkey(r[0]) in board_by_nkey})
    if variants:
        notes.append("case-only name variants reconciled by the normalised join key (not halted): "
                     + ", ".join(variants))
        check("case-only name variants reconciled by normalised key", True,
              "%d reconciled: %s" % (len(variants), ", ".join(variants)))

    # authored-ownership vs the stamped board (agreement report; board is the display membership source)
    def normt(t):
        return FREE_AGENTS if str(t).strip().lower() == "free agents" else str(t).strip()
    mismatch = []
    for r in loc_rows:
        bp = board_by_nkey.get(nkey(r[0]))
        if bp:
            ct, bt = normt(r[1]), bp[0]["affl_team"]
            if ct != bt:
                mismatch.append((r[0], ct, bt))
    check("authored CSV ownership agrees with the stamped board affl_team", not mismatch,
          "%d mismatch(es)%s" % (len(mismatch), ("" if not mismatch else ": " + str(mismatch[:5]))))
    if mismatch:
        notes.append("CSV ownership differs from the stamped board for %d player(s); the board rides "
                     "one bake behind for player club-moves (pick trades update on re-ingest). "
                     "First few: %s" % (len(mismatch), mismatch[:5]))
    return loc_enc_note(lenc, penc)


def loc_enc_note(lenc, penc):
    notes.append("input encodings: Player_Locations=%s · Future_Positioning=%s" % (lenc, penc))
    return True


# -------------------------------------------------------------------------- per-club valuation
def build_clubs(players, picks, affl_teams):
    picks_by_team = collections.defaultdict(list)
    for p in picks:
        picks_by_team[p["team"]].append(p)

    roster_by = collections.defaultdict(list)
    for p in players:
        t = p.get("affl_team")
        if t and t != FREE_AGENTS:
            roster_by[t].append(p)

    clubs = []
    for team in affl_teams:
        roster = sorted(roster_by.get(team, []), key=lambda x: -x["v"])
        total_player = sum(p["v"] for p in roster)
        top5 = sum(p["v"] for p in roster[:5])
        top10 = sum(p["v"] for p in roster[:10])

        # Best-23: greedy positional fill by highest board v per slot (CURRENT posCode), then best-5 bench.
        used = set()
        best23_keys = []
        best23 = 0
        for pos, n in SLOTS:
            picked = [p for p in roster if p["posCode"] == pos and id(p) not in used][:n]
            for p in picked:
                used.add(id(p)); best23 += p["v"]; best23_keys.append(p["key"])
        bench = [p for p in roster if id(p) not in used][:BENCH]
        for p in bench:
            used.add(id(p)); best23 += p["v"]; best23_keys.append(p["key"])

        tp = [t for t in picks_by_team[team]]
        total_picks = sum(p["value"] for p in tp)
        clubs.append({
            "team": team,
            "display": DISPLAY.get(team, team),
            "overall": total_player + total_picks,
            "totalPlayer": total_player,
            "totalPicks": total_picks,
            "top5": top5, "top10": top10,
            "best23": best23, "nonBest23": total_player - best23,
            "nRoster": len(roster), "nPicks": len(tp),
            "best23Keys": best23_keys,
        })
    clubs.sort(key=lambda c: -c["overall"])
    return clubs, picks_by_team


# ---------------------------------------------------------------------------------------- output
def _write(payload):
    body = ("// GENERATED by ui/tools/ingest_inputs.py — the no-LLM VALIDATE-OR-HALT club-valuation\n"
            "// ingest.  DO NOT hand-edit; regenerate after any docs/inputs/ change (see\n"
            "// ui/HOW_TO_UPDATE_INPUTS.md).  Pick prices are the engine's canonical PVC evaluated over\n"
            "// each band; no sheet value is ingested; no player value is recomputed.\n"
            "window.__CLUB_VALUATION__ = " + json.dumps(payload, ensure_ascii=False, sort_keys=True) + ";\n")
    with open(OUT, "w", encoding="utf-8") as f:
        f.write(body)


def _print_verdicts():
    print("\n  VALIDATION VERDICTS")
    print("  " + "-" * 72)
    for c, ok, d in verdicts:
        print("  [%s] %s%s" % ("PASS" if ok else "FAIL", c, ("  (%s)" % d) if d else ""))


# DISPLAY map mirrors ui/app/config.js CLUB_DISPLAY (display-only; long name stays the join key).
DISPLAY = {
    "North Melbourne Kangaroos": "North Melbourne",
    "Collingwood Magpies": "Collingwood",
    "Port Adelaide Power": "Port Adelaide",
}


def run():
    obj, stamp, pvc = load_board()
    players = obj.get("players", [])
    # Curve provenance: resolve the release-active pick curve from the explicit contract (deterministic,
    # fail-closed) BEFORE cross-checking the board PVC against it. Replaces the stale hardcoded L1b rule.
    contract = load_curve_contract()
    resolved = resolve_release_curve(contract)
    pvc = assert_pvc(pvc, resolved)

    board_by_nkey = collections.defaultdict(list)
    for p in players:
        board_by_nkey[nkey(p["name"])].append(p)
    check("board name-join keys unambiguous (normalised)",
          all(len(v) == 1 for v in board_by_nkey.values()),
          "dups: %s" % [k for k, v in board_by_nkey.items() if len(v) > 1])

    affl_teams = sorted({p["affl_team"] for p in players
                         if p.get("affl_team") and p["affl_team"] != FREE_AGENTS})
    check("AFFL club count", True, "%d clubs (+ Free Agents pool)" % len(affl_teams))

    validate_csvs(board_by_nkey)
    picks = load_picks(pvc, affl_teams)
    clubs, picks_by_team = build_clubs(players, picks, affl_teams)

    payload = {
        "stamp": {
            "board": stamp.get("board"), "engine": stamp.get("engine"), "store": stamp.get("store"),
            "tag": stamp.get("tag"), "expectedBoard": EXPECTED_BOARD, "baseYear": BASE_YEAR,
            "releaseVersion": contract.get("release_version"), "asOfRound": contract.get("as_of_round"),
            # Provenance of the release-active pick curve, resolved (never hardcoded) from the contract.
            "pvcSource": "%s (%s composed pathway; adopted %s; == rl_app_data PVC; pick1=3000)"
                         % (resolved["path"], resolved["gate"], contract.get("release_version")),
            "pvcPathway": resolved["gate"], "pvcCurveMd5": resolved["curve_md5"],
            "pvcCurveFileMd5": resolved["file_md5"], "releaseCurveContract": "ui/release_pick_curve.json",
            "pvcOk": True, "discount2027": PICK_FUTURE_DISCOUNT, "mult2027": 1.0 - PICK_FUTURE_DISCOUNT,
            "posture": "balanced (canonical; the only posture in this build)",
            "nPicks": len(picks), "nClubs": len(clubs), "generated": _now(),
        },
        "halt": None,
        "verdicts": [{"check": c, "ok": o, "detail": d} for c, o, d in verdicts],
        "notes": notes,
        "clubs": clubs,
        "picksByTeam": {t: [{"id": p["id"], "year": p["year"], "round": p["rnd"],
                             "origin": p["origin_team"], "originDisplay": DISPLAY.get(p["origin_team"], p["origin_team"]),
                             "low": p["lo"], "high": p["hi"], "band": p["band"], "value": p["value"]}
                            for p in sorted(picks_by_team[t], key=lambda x: (x["year"], -x["value"]))]
                        for t in picks_by_team},
    }
    _write(payload)
    _print_verdicts()
    print("\n  CLEAN INGEST — %d picks priced off %s (%s), %d clubs.  Bundle: ui/data/club_valuation.js"
          % (len(picks), resolved["path"], resolved["gate"], len(clubs)))
    print("\n  TOP-3 CLUBS BY OVERALL VALUE:")
    for c in clubs[:3]:
        print("    %-18s overall %s  (players %s + picks %s)" %
              (c["display"], f"{c['overall']:,}", f"{c['totalPlayer']:,}", f"{c['totalPicks']:,}"))
    return 0


def main():
    try:
        return run()
    except HaltError as e:
        _emit_halt(e.reason)
        return 2


if __name__ == "__main__":
    sys.exit(main())
