#!/usr/bin/env python3
"""
ITEM 262 — positional data landing + vocabulary replacement.

Committed as evidence per acceptance criterion 2: "Transcription is by script, never by hand;
the script and its inputs are committed as evidence."

Two stages, run separately so the proof can separate them (owner ruling Q1, 2026-07-29):

  stage1  vocabulary replacement ONLY. No owner edits, no new per-season data.
          The board must come out byte-identical modulo the relabel: zero value movers,
          zero rank movers.

  stage2  the owner's edits (43 present_position, 12 future_position, 1 alternate_position,
          2 p_dual_stream) + the new per-season eligibility column.
          Movers are EXPECTED here and every one is reported and attributed.

Usage:
    python3 migrate_positions.py stage1 [--check]
    python3 migrate_positions.py stage2 [--check]

--check writes nothing and reports what would change.

VOCABULARY (owner ruling 2026-07-29, issue #262 §4). Replaced, not merged, so any un-migrated
site fails visibly instead of half-matching:

    K-FWD / KFWD / KEY_FWD  -> KPF
    K-DEF / KDEF / KEY_DEF  -> KPD
    G-DEF / GDEF / GEN_DEF  -> SD
    G-FWD / GFWD / GEN_FWD  -> SF
    MID                     -> MID
    RUC  / RUCK             -> RUCK
    DEF                     -> SD      (Q3: the engine's existing rl_model.py:67
                                        GRP {'DEF':'GEN_DEF'} alias, made explicit)

Three spellings were live before this migration and collapse to one here:
  * concatenated  (GDEF)   — the store's *_position fields
  * hyphenated    (G-DEF)  — the store's `eligibilities` string, and the owner's CSV inputs
  * underscored   (GEN_DEF)— the engine's group names (rl_model.py GRP values)
After migration all three are the same six codes, so rl_model.py's GRP and _ELIG_MAP become
identity maps over the new vocabulary. That is intended, not an accident.
"""
import json, os, re, sys, hashlib, collections

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
SHEET = os.environ.get('ITEM262_SHEET',
    '/root/.claude/uploads/d4971030-9bc2-5f7f-83bc-62264dc5aff4/1317070f-rl_model_data_editable_1.xlsx')
STORE = os.path.join(REPO, 'engine/rl_after/rl_model_data.json')

# ---------------------------------------------------------------- vocabulary

NEW = {'KPF', 'KPD', 'SD', 'SF', 'MID', 'RUCK'}

# Longest-first so GEN_DEF is consumed before DEF, KFWD before FWD, RUCK before RUC.
TOKEN_MAP = [
    ('GEN_DEF', 'SD'),  ('GEN_FWD', 'SF'),  ('KEY_DEF', 'KPD'), ('KEY_FWD', 'KPF'),
    ('G-DEF',   'SD'),  ('G-FWD',   'SF'),  ('K-DEF',   'KPD'), ('K-FWD',   'KPF'),
    ('GDEF',    'SD'),  ('GFWD',    'SF'),  ('KDEF',    'KPD'), ('KFWD',    'KPF'),
    ('RUCK',    'RUCK'),('RUC',     'RUCK'),
    ('MID',     'MID'),
    ('DEF',     'SD'),  # Q3 alias — only ever reached after the compound forms above
]
_ALT = '|'.join(re.escape(k) for k, _ in TOKEN_MAP)
_TOKEN_RE = re.compile(r'(?<![A-Za-z0-9_])(' + _ALT + r')(?![A-Za-z0-9_])')
_LOOKUP = dict(TOKEN_MAP)

# For JSON files: match a token ONLY when it is a complete quoted string (a dict key or a
# bare-token value), optionally behind a '<prefix>|' composite as in rl_passmark's "3|GEN_DEF".
# This deliberately does NOT match tokens inside prose. ycred_table.json carries a 2KB `note`
# describing the L1c derivation that mentions GEN_DEF and RUC; rewriting it would falsify a
# historical record, and the structural cross-check below fails loudly if the two ever disagree.
_JSON_TOKEN_RE = re.compile(r'"((?:[^"\\]*\|)?)(' + _ALT + r')"')


def remap_token(tok):
    """Map one bare position token. Raises on anything unrecognised — never guesses."""
    if tok in NEW:
        return tok
    if tok not in _LOOKUP:
        raise KeyError('unmapped position token %r' % tok)
    return _LOOKUP[tok]


def remap_text(s):
    """Replace every position token in a string, leaving all other text byte-identical."""
    return _TOKEN_RE.sub(lambda m: _LOOKUP[m.group(1)], s)


def remap_csv_field(s):
    """A comma-separated eligibility string, e.g. 'MID,G-FWD' -> 'MID,SF'. Order preserved."""
    if s is None:
        return None
    parts = [p.strip() for p in str(s).split(',') if p.strip()]
    out, seen = [], set()
    for p in parts:
        v = remap_token(p)
        if v not in seen:          # K-X absorbs G-X can produce a duplicate; keep first
            seen.add(v); out.append(v)
    return ','.join(out)


# ------------------------------------------------------- file classification
#
# Measured, not assumed: the live-input set is every file the board build actually OPENS
# (audit-hook trace of `python3 rl_export.py`, 998 opens) that also carries an old token.
# Everything else carrying a token is historical and is enumerated as exempt below.

STORE_POSITION_FIELDS = ('drafted_position', 'present_position', 'future_position',
                         'alternate_position')

# The one prose field in the store carrying old vocabulary. Owner ruling R2-4 (#262 Addendum 4):
# rewrite it in the rename pass — "The note exists only to keep Aiden distinct from Nathan
# O'Driscoll; the relabel doesn't touch that purpose." With this edit the store carries zero
# old-vocabulary hits and needs no enumerated exemption.
# Keyed by player `key` and matched in full, so it cannot silently hit a different row.
STORE_PROSE_REWRITES = {
    'aiden-o-driscoll': {
        'pick_correction_note': (
            'Historical position corrected to GDEF by owner ruling.',
            'Historical position corrected to SD by owner ruling.'),
    },
}

# JSON files whose position tokens are dispatch keys/values the engine reads.
JSON_DATA_FILES = [
    'engine/rl_after/params.json',
    'engine/rl_after/rl_passmark.json',
    'engine/rl_after/bust_prior_table.json',
    'engine/rl_after/lti_return_table.json',
    'engine/rl_after/ycred_table.json',
    # A LIVE PINNED EXPECTATION, not a frozen record. The collision sentry pins the drafted
    # position of the two separately-keyed Max Kings so an identity bleed between them is
    # caught. Its purpose is telling them apart, which the relabel does not touch — so the
    # pinned values must move with the store or the sentry reds on the rename itself.
    # one_source_selftest caught this; the build trace could not, because it is a test.
    'engine/rl_after/collision_sentry.json',
]

# Python/JS sources whose position tokens are literals the engine dispatches on.
#
# pgrid.py earns its place the expensive way. It was MISSING from the first live-input set
# because that set was measured with an audit hook listening only for `open` — which never
# fires for a module pulled in through the import machinery. pgrid is imported by rl_model,
# so it was invisible. It also carries the second half of a DUPLICATED grp3(): rl_model.py:954
# and pgrid.py:55 independently compute the same coarse GEN/KEY/RUC bucket, one READING the
# grid the other BUILDS. Migrating one side and not the other produced KeyError 'RUCK' at
# pgrid.Praw — hazard 2 (duplicated assertion) meeting hazard 7 (build-on-one-axis,
# read-on-the-other). The replaced-vocabulary design is what made it fail loudly instead of
# half-matching; a merged vocabulary would have silently bucketed every ruck as GEN.
CODE_FILES = [
    'engine/rl_after/rl_model.py',
    'engine/rl_after/pgrid.py',
    'engine/rl_after/_merged_recover.py',
    'engine/rl_after/rl_export.py',
    'engine/forward_valuation/build_cohort_book.py',
    'engine/forward_valuation/build_peak_model_v4.py',
    'engine/forward_valuation/conditional_prior.py',
    'engine/forward_valuation/dist_redesign.py',
    'engine/forward_valuation/par_build.py',
    'engine/forward_valuation/par_redesign.py',
    'engine/forward_valuation/tail_restore.py',
    # TESTS AND GATES — acceptance criterion 2 names "store, engine, board data, UI, tests".
    # None of these appear in a board-build trace because they are the harness, not the build;
    # they were added after one_source_selftest.py failed on its own old-spelling fixtures.
    'engine/rl_after/one_source_selftest.py',
    'engine/rl_after/verify_anchors.py',
    'ship_gates_check.py',
    # further live engine sources carrying position literals
    'engine/rl_after/_flags_support.py',
    'engine/rl_after/_gate1.py',
    'engine/rl_after/_gate1_wf.py',
    'engine/rl_after/_p2b_divergence.py',
    'engine/rl_after/_comb_book.py',
    'engine/rl_after/_build_book_xlsx.py',
    'engine/rl_after/_langdon_decomp.py',
    'engine/rl_after/_ov_angleB.py',
    'engine/rl_after/_engine_block_v23.js',
    'engine/rl_after/derive_lti_return.py',
    'engine/rl_after/s4_matrix_M1v7_blend.py',
    'engine/rl_after/s4_matrix_M1v7_retainonly.py',
    'engine/rl_after/ingestion/round_movers.py',
]

# The two translation tables in rl_model.py collapse to identity maps under the rename,
# because the three spellings they used to translate between are now one vocabulary. A blind
# token replace would leave DUPLICATE dict keys behind ('GDEF' and the 'DEF' alias both map to
# 'SD'; 'RUC' and 'RUCK' both to 'RUCK') — legal Python, but sloppy in source and it hides the
# fact that the DEF alias is now dead. These two lines are rewritten explicitly instead.
#
# Each pattern must match EXACTLY ONCE or the migration halts (anchoring sentinel, hazard 14).
SPECIAL_REWRITES = {
    'engine/rl_after/rl_model.py': [
        (r"^GRP=\{'MID':'MID','RUC':'RUC','GFWD':'GEN_FWD','KFWD':'KEY_FWD',"
         r"'GDEF':'GEN_DEF','DEF':'GEN_DEF','KDEF':'KEY_DEF'\}",
         "GRP={'MID':'MID','RUCK':'RUCK','SF':'SF','KPF':'KPF','SD':'SD','KPD':'KPD'}"
         "   # ITEM 262: identity since the store now speaks the engine's own vocabulary."
         " Kept as the dispatch boundary, NOT redundant. The pre-262 'DEF'->'GEN_DEF'"
         " back-catalogue alias is retired: all 136 rows were migrated to SD."),
        (r"^_ELIG_MAP=\{'MID':'MID','RUC':'RUC','RUCK':'RUC','G-FWD':'GEN_FWD',"
         r"'K-FWD':'KEY_FWD','G-DEF':'GEN_DEF','K-DEF':'KEY_DEF'\}",
         "_ELIG_MAP={'MID':'MID','RUCK':'RUCK','SF':'SF','KPF':'KPF','SD':'SD','KPD':'KPD'}"
         "   # ITEM 262: identity — `eligibilities` no longer uses a separate hyphenated"
         " spelling. Kept so _collapse_elig still validates its input."),
    ],
}

# EXEMPT AND ENUMERATED (acceptance criterion 2 requires the list, not a blanket).
#
#  (a) prose notes describing past events. Renaming them would falsify the historical
#      record — the note says what the vocabulary WAS at the time it was written.
EXEMPT_PROSE = {
    'data/expected_boot.json':        ['_engine_head_note', '_legc_note',
                                       '_legd_act2_note', '_r1067_note'],
    'data/model_config.json':         ['var_notes.RL_PVCADOPT'],
    'engine/rl_after/pvc_curve_L1b.json': ['note'],
    'engine/rl_after/ycred_table.json':   ['note'],   # 2KB L1c derivation narrative
}
#  (b) hash-named frozen snapshots. The hash in the FILENAME identifies the engine head or
#      config that produced the contents; rewriting the contents would make the name a lie.
#      This is hazard class 1 (right-name-wrong-file) and the rename must not touch them.
EXEMPT_FROZEN_GLOBS = [
    'data/gates_snapshots/gates_*.json',       # 23 files
    'data/s4_matrix_baked_*.json',
    'data/s4_matrix_control_*.json',
    'data/s4_matrix_gradedfix_*.json',
    'data/s4_matrix_v2*_*.json',
    'data/s4_matrix_v2_*.json',
    'data/s4_matrix_nogames.json',
    'engine/rl_after/ingestion/movers/movers_R*.json',
    'engine/rl_after/s4_matrix_M1v7*.json',
]
#  (c) archives and the register — exempt by directive. The register's exemption extends to its
#      new home (docs/register/, the 3b act 2026-08-21): the durable record's prose is never
#      rewritten by a token sweep, in either of its two homes.
EXEMPT_TREES = ('session_', 'backups/', 'evidence/', 'docs/archive/', 'docs/register/')
EXEMPT_FILES = ('docs/OPEN_ITEMS_REGISTER.md', 'LTI_REGISTER.md')
#  (d) regenerated, never hand-edited — the board is rebuilt by rl_export.py.
REGENERATED = ('data/rl_build/rl_app_data.json',)


# ------------------------------------------------------------------ stage 1

def json_remap_keys_and_values(obj, touched):
    """Walk a JSON structure remapping bare position tokens in dict KEYS and in string
    VALUES that are wholly position tokens (or '<n>|<TOKEN>' composites). Numbers, prose
    and everything else are left byte-identical."""
    if isinstance(obj, dict):
        out = {}
        for k, v in obj.items():
            nk = k
            if isinstance(k, str):
                if '|' in k:
                    head, _, tail = k.rpartition('|')
                    if tail in _LOOKUP or tail in NEW:
                        nk = head + '|' + remap_token(tail);
                elif k in _LOOKUP or k in NEW:
                    nk = remap_token(k)
            if nk != k:
                touched.append(('key', k, nk))
            out[nk] = json_remap_keys_and_values(v, touched)
        return out
    if isinstance(obj, list):
        return [json_remap_keys_and_values(v, touched) for v in obj]
    if isinstance(obj, str) and (obj in _LOOKUP or obj in NEW):
        nv = remap_token(obj)
        if nv != obj:
            touched.append(('val', obj, nv))
        return nv
    return obj


def stage1_store(check):
    """Rename the store's own position vocabulary. Nothing else in the store moves."""
    with open(STORE) as f:
        players = json.load(f)
    counts = collections.Counter()
    for p in players:
        for fld in STORE_POSITION_FIELDS:
            if p.get(fld):
                old = p[fld]
                p[fld] = remap_token(old)
                if p[fld] != old:
                    counts[(fld, old, p[fld])] += 1
        if p.get('eligibilities'):
            old = p['eligibilities']
            p['eligibilities'] = remap_csv_field(old)
            if p['eligibilities'] != old:
                counts[('eligibilities', old, p['eligibilities'])] += 1
        for fld, (was, now) in STORE_PROSE_REWRITES.get(p.get('key'), {}).items():
            if p.get(fld) == was:
                p[fld] = now
                counts[(fld + ' [R2-4 prose]', was, now)] += 1
            elif p.get(fld) == now:
                pass                      # already applied; the pass is idempotent
            else:
                raise SystemExit('HALT: %s.%s is %r, not the text R2-4 rules on — stop and '
                                 'look' % (p.get('key'), fld, p.get(fld)))
    if not check:
        # PROVEN byte-exact: json.dumps(obj) on the UNMODIFIED store reproduces the file
        # md5 e3aaba77 exactly, so this encoder is the one that wrote it. Anything other
        # than the position fields above is therefore byte-identical by construction.
        with open(STORE, 'w') as f:
            json.dump(players, f)
    return counts, len(players)


def stage1_files(check):
    """The small JSON tables carry position tokens ONLY as dict keys (verified by walking
    each file), so they are migrated by format-preserving TEXTUAL replacement rather than a
    JSON round-trip. params.json uses a hand-rolled indent that no json.dumps setting
    reproduces byte-exact; a structural rewrite would reformat all 1,689 bytes of it and
    bury the 15 real changes."""
    report = {}
    for rel in JSON_DATA_FILES + CODE_FILES:
        path = os.path.join(REPO, rel)
        with open(path) as f:
            src = f.read()
        if rel.endswith('.json'):
            changed = sum(1 for m in _JSON_TOKEN_RE.finditer(src)
                          if _LOOKUP[m.group(2)] != m.group(2))
            new = _JSON_TOKEN_RE.sub(
                lambda m: '"%s%s"' % (m.group(1), _LOOKUP[m.group(2)]), src)
            a, b = json.loads(src), json.loads(new)
            if json.dumps(json_remap_keys_and_values(a, []), sort_keys=True) != \
               json.dumps(b, sort_keys=True):
                raise SystemExit('HALT: textual migration of %s does not match the '
                                 'structural migration — stop and look' % rel)
            if not check and new != src:
                with open(path, 'w') as f:
                    f.write(new)
            report[rel] = changed
            continue
        changed = sum(1 for m in _TOKEN_RE.finditer(src) if _LOOKUP[m.group(1)] != m.group(1))
        # explicit rewrites first, so the blind pass never sees these lines
        for pat, repl in SPECIAL_REWRITES.get(rel, []):
            hits = len(re.findall(pat, src, flags=re.M))
            if hits != 1:
                raise SystemExit('HALT: %s special rewrite expected exactly 1 match, '
                                 'found %d for %r' % (rel, hits, pat[:60]))
            src = re.sub(pat, lambda _m: repl, src, flags=re.M)
        new = remap_text(src)
        if SPECIAL_REWRITES.get(rel):
            for _pat, repl in SPECIAL_REWRITES[rel]:
                if repl.split('   #')[0] not in new:
                    raise SystemExit('HALT: %s special rewrite did not survive the '
                                     'token pass' % rel)
        if not check and new != src:
            with open(path, 'w') as f:
                f.write(new)
        report[rel] = changed
    return report


# ------------------------------------------------------------------ stage 2
#
# The Q5/Q6 per-season table, owner ruling 2026-07-29 + Addendum 3. "KEY until year X" means up
# to AND INCLUDING X. Entries written "when FWD/DEF eligible" key off the season's own eligibility
# rather than a year range.
#
# Each entry answers ONE question: for this player, in this season, is this FWD-or-DEF component
# key or general? Returning None means "not named — fall through to the standing scope rule".
_ALL = range(1900, 2100)

def _years(*ys):
    s = set()
    for y in ys:
        s.update(y if isinstance(y, (range, list, tuple, set)) else [y])
    return s

# name -> callable(year, comp) -> True (key) / False (general)
Q56_TABLE = {
    # position-conditional, not year-conditional
    'Nick Blakey':        lambda y, c: c == 'FWD',          # FWD -> KPF, DEF -> SD
    'Harrison Himmelberg':lambda y, c: c == 'FWD',          # R2-2: "Harry" = Harrison, src 1107
    'Mitch McGovern':     lambda y, c: c == 'FWD',
    'Jayden Laverde':     lambda y, c: c == 'DEF',          # FWD -> SF, DEF -> KPD
    'Mark Blicavs':       lambda y, c: c == 'DEF',          # KEY when DEF eligible
    'Andy Otten':         lambda y, c: c == 'DEF',          # KEY when DEF
    # year-ranged, both FWD and DEF key in the named years
    'James Sicily':       lambda y, c: y in _years(2020, range(2023, 2026)),
    'Jordan Ridley':      lambda y, c: y in _years(range(2022, 2026)),
    'Josh Worrell':       lambda y, c: y in _years(range(2024, 2026)),
    'Jack Scrimshaw':     lambda y, c: y == 2025,
    'Jake Melksham':      lambda y, c: y in _years(2025, 2026),
    'Dane Rampe':         lambda y, c: y in _years(range(2017, 2025)),
    'Mitchito Owens':     lambda y, c: y in _years(2024, 2026),
    'Nick Haynes':        lambda y, c: y <= 2017,
    'Mason Wood':         lambda y, c: y <= 2015,
    'Jeremy Howe':        lambda y, c: y in _years(range(2020, 2026)),
    'Kyle Langford':      lambda y, c: y in _years(2025, 2026),
    'Reef McInnes':       lambda y, c: True,                # KEY for his whole career
    'Ryan Lester':        lambda y, c: y in _years(2024, 2025),
    'Jack Lukosius':      lambda y, c: y <= 2019 or y in _years(2023, 2024, 2026),
    'Jai Serong':         lambda y, c: y <= 2026,
    'Jake Lever':         lambda y, c: y <= 2025,
    'Jy Farrar':          lambda y, c: y >= 2025,
    # explicitly no key seasons
    'Reuben Ginbey':      lambda y, c: False,               # key from FUTURE seasons only
    'Ryan Maric':         lambda y, c: False,               # drafted key, never played it
}

# Addendum 3 R2-1: position-field overrides that differ from BOTH the store and the sheet.
# Jake Lever is deliberately absent — his fields are already key-coded, so he contributes zero
# stage-2 edits; his future-key status lives in the position fields, not in new season rows.
R2_1_FIELD_OVERRIDES = {
    'Ryan Lester':   {'present_position': 'SD',  'future_position': 'SD'},
    'Jack Lukosius': {'present_position': 'SF',  'future_position': 'SF'},
}

# The sheet's explicit per-season spellings (Harrison Himmelberg only) map straight through.
SHEET_EXPLICIT = {'KEY FWD': 'KPF', 'GEN DEF': 'SD', 'KEY DEF': 'KPD', 'GEN FWD': 'SF'}

# Q4: any DPP season containing RUCK renders its FWD/DEF component key, regardless of flag.
# SF/RUCK and SD/RUCK are not possible. Pure RUCK has no key variant.
_SEASON_BASE = {'FWD', 'DEF', 'MID', 'RUCK'}


def resolve_season(sheet_pos, player_name, is_key, year):
    """One season's sheet Position -> the landed per-season eligibility, in the new vocabulary.
    Raises on anything it does not recognise — it never guesses."""
    if sheet_pos in SHEET_EXPLICIT:                      # rule 1: transcribe as given
        return SHEET_EXPLICIT[sheet_pos], 'sheet-explicit'
    parts = [p.strip() for p in str(sheet_pos).split('/')]
    for p in parts:
        if p not in _SEASON_BASE:
            raise KeyError('unrecognised season position %r for %s %s'
                           % (sheet_pos, player_name, year))
    has_ruck = 'RUCK' in parts
    out, why = [], set()
    for c in parts:                                      # rule 5: sheet order preserved
        if c == 'MID':
            out.append('MID'); continue
        if c == 'RUCK':
            out.append('RUCK'); continue
        if has_ruck:                                     # rule 3a: Q4 override wins
            key, w = True, 'Q4-ruck'
        elif player_name in Q56_TABLE:                   # rule 3b: the named table
            key, w = bool(Q56_TABLE[player_name](year, c)), 'Q5/Q6-table'
        elif is_key:                                     # rule 3c: blanket overlay
            key, w = True, 'blanket'
        else:
            key, w = False, 'general'
        out.append(('KPF' if c == 'FWD' else 'KPD') if key else ('SF' if c == 'FWD' else 'SD'))
        why.add(w)
    return '/'.join(out), '+'.join(sorted(why)) or 'plain'


def _load_sheet():
    import openpyxl, warnings
    warnings.filterwarnings('ignore')
    wb = openpyxl.load_workbook(SHEET, read_only=True, data_only=True)
    it = wb['Players'].iter_rows(values_only=True); ph = list(next(it))
    players = [dict(zip(ph, r)) for r in it]
    players = [r for r in players if any(v is not None for v in r.values())]
    it = wb['Scoring History'].iter_rows(values_only=True); sh = list(next(it))
    seasons = [dict(zip(sh, r)) for r in it]
    by = collections.defaultdict(list)
    for s in seasons:
        by[s['source_row']].append(s)
    return players, by


# Player-level fields the owner edited in the sheet. affl_team is DELIBERATELY ABSENT:
# owner ruling Q2 routes those 16 trades through the #232 ownership sidecar, not the store.
SHEET_EDIT_FIELDS = ('present_position', 'future_position', 'alternate_position', 'p_dual_stream')


def stage2(check):
    sheet, seasons_by_row = _load_sheet()
    with open(STORE) as f:
        store = json.load(f)
    if len(sheet) != len(store):
        raise SystemExit('HALT: sheet %d rows vs store %d' % (len(sheet), len(store)))
    # The Players tab is NOT in source_row order — Shiel/Cameron/Treloar (the notional-pick
    # trio, Addendum 2) sit above row 1. Index by source_row, never by sheet position.
    sheet_by_row = {r['source_row']: r for r in sheet}
    if len(sheet_by_row) != len(sheet):
        raise SystemExit('HALT: duplicate source_row in the sheet')

    edits, season_stats, why_stats = [], collections.Counter(), collections.Counter()
    bust_default = []

    for srow, p in enumerate(store, start=1):
        s = sheet_by_row.get(srow)
        if s is None:
            raise SystemExit('HALT: store row %d has no sheet row' % srow)
        if s['player'] != p['player']:
            raise SystemExit('HALT: row %d name mismatch %r vs %r'
                             % (srow, s['player'], p['player']))

        # ---- (a) the owner's deliberate player-level edits, mapped through the vocabulary
        for fld in SHEET_EDIT_FIELDS:
            want = s.get(fld)
            if fld != 'p_dual_stream' and want:
                want = remap_token(want)
            if fld == 'p_dual_stream' and want is not None:
                want = int(want)
            have = p.get(fld)
            if have != want and not (have is None and want is None):
                edits.append((srow, p['player'], fld, have, want, 'sheet'))
                if not check:
                    p[fld] = want

        # ---- (b) Addendum 3 R2-1 overrides, which differ from BOTH store and sheet
        for fld, want in R2_1_FIELD_OVERRIDES.get(p['player'], {}).items():
            if p.get(fld) != want:
                edits.append((srow, p['player'], fld, p.get(fld), want, 'R2-1'))
                if not check:
                    p[fld] = want

        # ---- (c) the new per-season eligibility
        rows = seasons_by_row.get(srow, [])
        if len(rows) != len(p.get('scoring') or []):
            raise SystemExit('HALT: %s has %d sheet seasons vs %d store seasons'
                             % (p['player'], len(rows), len(p.get('scoring') or [])))
        byyear = {r['year']: r for r in rows}
        for rec in (p.get('scoring') or []):
            sr = byyear.get(rec['year'])
            if sr is None:
                # the ruled bust default: no training-data season -> drafted position, MARKED.
                # Owner ruling Q9 says this is vacuous; the branch exists so a future season
                # without sheet data is marked rather than silently defaulted.
                rec['pos'] = p['drafted_position']
                rec['pos_src'] = 'drafted-default'
                bust_default.append((p['player'], rec['year']))
                continue
            landed, why = resolve_season(sr['Position'], p['player'],
                                         s.get('is_key?') == 'KEY', rec['year'])
            if not check:
                rec['pos'] = landed
            season_stats[landed] += 1
            why_stats[why] += 1

    if not check:
        with open(STORE, 'w') as f:
            json.dump(store, f)

    print('\n-- (a)+(b) player-level edits: %d --' % len(edits))
    bysrc = collections.Counter(e[5] for e in edits)
    byfld = collections.Counter(e[2] for e in edits)
    print('   by source:', dict(bysrc), '  by field:', dict(byfld))
    for srow, nm, fld, a, b, src in edits:
        if src == 'R2-1':
            print('   [R2-1] src=%-5d %-22s %-20s %s -> %s' % (srow, nm, fld, a, b))
    print('\n-- (c) per-season eligibility: %d seasons --' % sum(season_stats.values()))
    for k, v in sorted(season_stats.items(), key=lambda x: -x[1]):
        print('   %-14s %6d' % (k, v))
    print('\n   resolution path:')
    for k, v in sorted(why_stats.items(), key=lambda x: -x[1]):
        print('   %-22s %6d' % (k, v))
    print('\n   bust-default seasons applied: %d %s'
          % (len(bust_default), '(Q9: ruled vacuous)' if not bust_default else bust_default[:5]))
    if not check:
        print('\nstore md5 after: %s' % hashlib.md5(open(STORE,'rb').read()).hexdigest())
    return 0


def revert62(check):
    """Owner ruling R3-1: the 62 position-field edits DEFER to the re-derivation job, which
    applies them as its first act with every mover attributed there. This amends Addendum 2 Q1's
    "land them" by owner word.

    Restores the four edited player-level fields to their stage-1 (post-rename, pre-edit) values,
    taken from the committed stage-1 store, and keeps everything stage 2 added that stays:
    the per-season `pos` eligibility and the R2-4 O'Driscoll note.

    The stage-2 write is deliberately NOT rewritten out of history — history is the record."""
    import subprocess
    ref = os.environ.get('ITEM262_STAGE1_REF', 'HEAD~1')
    raw = subprocess.run(['git', '-C', REPO, 'show',
                          '%s:engine/rl_after/rl_model_data.json' % ref],
                         capture_output=True, text=True)
    if raw.returncode != 0:
        raise SystemExit('HALT: cannot read the stage-1 store at %s' % ref)
    stage1 = {p['key']: p for p in json.loads(raw.stdout)}
    with open(STORE) as f:
        store = json.load(f)
    if len(stage1) != len(store):
        raise SystemExit('HALT: stage-1 store %d records vs current %d'
                         % (len(stage1), len(store)))
    reverted = collections.Counter(); rows = []
    kept_pos = 0
    for p in store:
        s1 = stage1.get(p['key'])
        if s1 is None:
            raise SystemExit('HALT: %s absent from the stage-1 store' % p['key'])
        for fld in SHEET_EDIT_FIELDS:
            if p.get(fld) != s1.get(fld):
                rows.append((p['player'], fld, p.get(fld), s1.get(fld)))
                reverted[fld] += 1
                if not check:
                    if fld in s1: p[fld] = s1[fld]
                    else: p.pop(fld, None)
        kept_pos += sum(1 for r in (p.get('scoring') or []) if 'pos' in r)
    if not check:
        with open(STORE, 'w') as f:
            json.dump(store, f)
    print('\n-- R3-1 revert: %d field values restored to stage 1 --' % sum(reverted.values()))
    print('   by field:', dict(reverted))
    print('   per-season `pos` keys KEPT: %d' % kept_pos)
    if not check:
        print('   store md5 after revert: %s'
              % hashlib.md5(open(STORE,'rb').read()).hexdigest())
    return 0


# ------------------------------------------------------------------- driver

def main():
    if len(sys.argv) < 2 or sys.argv[1] not in ('stage1', 'stage2', 'revert62'):
        print(__doc__); return 2
    stage, check = sys.argv[1], '--check' in sys.argv
    print('ITEM 262 %s%s' % (stage, '  [--check, writing nothing]' if check else ''))
    print('repo   : %s' % REPO)
    print('store  : %s  md5 %s' % (STORE, hashlib.md5(open(STORE,'rb').read()).hexdigest()))
    if stage == 'stage1':
        counts, n = stage1_store(check)
        print('\n-- store: %d player records --' % n)
        for (fld, old, new), c in sorted(counts.items(), key=lambda x: (-x[1], x[0])):
            print('   %-20s %-14s -> %-14s %5d' % (fld, old, new, c))
        print('   total field values remapped: %d' % sum(counts.values()))
        rep = stage1_files(check)
        print('\n-- engine data + code --')
        for rel, c in rep.items():
            print('   %-52s %5d tokens' % (rel, c))
        print('\nEXEMPT, enumerated:')
        print('   prose notes            : %d fields across %d files'
              % (sum(len(v) for v in EXEMPT_PROSE.values()), len(EXEMPT_PROSE)))
        print('   hash-named snapshots   : %d glob patterns (contents frozen; renaming '
              'would falsify the filename hash — hazard 1)' % len(EXEMPT_FROZEN_GLOBS))
        print('   archives + register    : %s %s' % (EXEMPT_TREES, EXEMPT_FILES))
        print('   regenerated by build   : %s' % (REGENERATED,))
        if not check:
            print('\nstore md5 after: %s'
                  % hashlib.md5(open(STORE,'rb').read()).hexdigest())
    elif stage == 'stage2':
        return stage2(check)
    else:
        return revert62(check)
    return 0


if __name__ == '__main__':
    sys.exit(main())
