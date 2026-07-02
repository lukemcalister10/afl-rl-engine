import io, contextlib, numpy as np
g={}
with contextlib.redirect_stdout(io.StringIO()):
    exec(open('_merged_recover.py').read().split('print("=== AFTER')[0], g)
MA=g['MA']; cp=g['cp']; dp=g['dp']; rd=g['rd']; PR=g['PR']; b6=g['b6']; WQ6=g['WQ6']; era=g['era']; REF=g['REF']
_lvlcurr=g['_lvlcurr']; _nqual=g['_nqual']; PROVEN_N=g['PROVEN_N']; DOWN_TOL=g['DOWN_TOL']
_agemult=g['_agemult']; _par_prior=g['_par_prior']; _upS=g['_upS']; _eo=g['_eo']; delisted=g['delisted']
def adj(a,y): return a*REF/era.get(y,REF)
TOL_M1=5.0; G_ADQ=12; WIN=2; S_M1=0.46; GCAP=17.0
def _radq(p,Y,Lo): return any(x['games']>=G_ADQ and x['avg']>Lo for x in p['scoring'] if Y-WIN<x['year']<=Y and (cp.debutyr(p)-1)<x['year'])
def _core(p,Y):
    Lo=cp._lvl_eff_orig(p,Y); n=_nqual(p,Y)
    if n==0: return Lo
    Lc=_lvlcurr(p,Y)
    if n>=PROVEN_N:
        if Lc>=Lo: return (Lo+S_M1*(Lc-Lo)) if ((Lc-Lo)>=TOL_M1 and _radq(p,Y,Lo)) else Lo
        drop=Lo-Lc
        if drop<=DOWN_TOL: return Lo
        sw=float(np.clip((drop-DOWN_TOL)/5,0,1)); return (1-sw)*Lo+sw*Lc*_agemult(cp._age_asof(p,Y))
    c=n/PROVEN_N; return c*Lc+(1-c)*_par_prior(p,Y)
def _infer(p,Y):
    L0=_core(p,Y); eo=_eo(p,Y)
    if eo<=0: return L0
    avs=[x['avg'] for x in p['scoring'] if x.get('games',0)>=6 and (cp.debutyr(p)-1)<x['year']<=Y]
    if not avs: return L0
    bar=MA.REPL.get(MA.gfut(p),0.0)-3.0; N=Y-cp.debutyr(p)+1
    return (1-eo)*L0+eo*min(L0,max(_upS(max(avs)-bar,N),_lvlcurr(p,Y)))
ORIG=cp._lvl_eff
def eff_s(p,Y): return sum(min(x['games']/GCAP,1.0) for x in p['scoring'] if x['games']>=6 and (cp.debutyr(p)-1)<x['year']<=Y)
def band_at(p,Y,fn):
    o=cp._lvl_eff; cp._lvl_eff=fn
    try: MA._pe_clear(); bb=list(b6(p,Y))
    finally: cp._lvl_eff=o; MA._pe_clear()
    return bb
def price(p,Y,bb,w):
    w=np.array(w)/sum(w); sav=dict(MA.REPL)
    try:
        for gg in MA.REPL: MA.REPL[gg]=sav[gg]-rd.REPL_DROP.get(gg,0)
        MA.BASE_REF=MA.AGE_REF=Y; MA._pe_clear()
        with contextlib.redirect_stdout(io.StringIO()): v=[dp.v_at_peak(p,float(L),'bal') for L in bb]
        return float(dp.SCALE_DIST*np.dot(w,v))
    finally: MA.REPL.update(sav)
def v7c(bb,p,Y):
    bb=list(bb); m=bb[2]; a=cp._age_asof(p,Y); cB=0.47*float(np.clip((eff_s(p,Y)-1)/3,0,1))
    asc=float(np.interp(a,[20,22,24,27],[1.0,0.76,0.58,0.40]))
    bb[3]=m+(1-cB)*(bb[3]-m); bb[4]=m+(1-cB)*(bb[4]-m); bb[5]=m+asc*(bb[5]-m); return bb
def qs(p,Y): return sorted([(x['year'],x['games'],adj(x['avg'],x['year'])) for x in p['scoring'] if x['games']>=6 and x['year']<=Y])
def trend(p,Y):
    l3=[s[2] for s in qs(p,Y)][-3:]; return (l3[-1]-l3[0]) if len(l3)>=2 else 0.0
def meang(p,Y):
    s=qs(p,Y); return float(np.mean([x[1] for x in s])) if s else 0.0
Y=2026; POS6=('MID','GEN_DEF','GEN_FWD','KEY_DEF','KEY_FWD','RUC'); rows=[]
for p in MA.data:
    if MA.gfut(p) not in POS6 or delisted(p) or g['nseas'](p,Y)<2: continue   # v7 applies to EXPERIENCED players only
    try: cur=price(p,Y,band_at(p,Y,ORIG),WQ6)
    except Exception: continue
    if cur<200: continue
    comb=price(p,Y,v7c(band_at(p,Y,_infer),p,Y),WQ6)
    Lo=cp._lvl_eff_orig(p,Y); Lc=_lvlcurr(p,Y); n=_nqual(p,Y); N=Y-cp.debutyr(p)+1; mg=meang(p,Y); tr=trend(p,Y)
    fired=(n>=PROVEN_N and Lc>=Lo and (Lc-Lo)>=TOL_M1 and _radq(p,Y,Lo))
    if N<=1: arch='first-year'
    elif mg<12 and n>=3: arch='injury-prone'
    elif (Lc-Lo)<=-5 or tr<=-8: arch='decliner'
    elif (Lc-Lo)>=5: arch='rising'
    else: arch='plateau'
    tier='elite' if cur>=2500 else 'upper' if cur>=1500 else 'mid' if cur>=700 else 'low'
    note='trend-up/level-flat (Flag A)' if (n>=PROVEN_N and tr>=8 and (Lc-Lo)<5) else ''
    rows.append([p['player'],MA.gfut(p),round(cp._age_asof(p,Y)),n,N,round(mg,1),round(tr,1),round(Lc-Lo,1),
                 round(cur),round(comb),(comb-cur)/cur,tier,arch,'Y' if fired else '',note])
rows.sort(key=lambda r:r[10])   # by delta% ascending (crushes first)
print("rows:",len(rows))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
wb=Workbook(); F='Arial'
hdr_fill=PatternFill('solid',start_color='1F3864'); hdr_font=Font(name=F,bold=True,color='FFFFFF',size=10)
thin=Side(style='thin',color='D9D9D9')

# --- Book sheet ---
ws=wb.active; ws.title='Book'
cols=['Player','Pos','Age','QualSeas','YrsSince','MeanG','Trend','LvlGap','Current','Combined','Delta%','Tier','Archetype','M1fired','Note']
ws.append(cols)
for r in rows: ws.append(r)
for j,c in enumerate(cols,1):
    cell=ws.cell(1,j); cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
widths=[22,9,5,9,9,7,7,7,9,9,8,7,13,8,18]
for j,wd in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=wd
nrow=len(rows)+1
for i in range(2,nrow+1):
    for j in range(1,len(cols)+1):
        cl=ws.cell(i,j); cl.font=Font(name=F,size=9); cl.border=Border(bottom=thin)
    ws.cell(i,9).number_format='#,##0'; ws.cell(i,10).number_format='#,##0'; ws.cell(i,11).number_format='0.0%'
    ws.cell(i,6).number_format='0.0'; ws.cell(i,7).number_format='+0.0;-0.0'; ws.cell(i,8).number_format='+0.0;-0.0'
    if ws.cell(i,14).value=='Y': ws.cell(i,14).font=Font(name=F,size=9,bold=True,color='006100')
ws.freeze_panes='A2'; ws.auto_filter.ref=f"A1:{get_column_letter(len(cols))}{nrow}"
ws.conditional_formatting.add(f"K2:K{nrow}", ColorScaleRule(
    start_type='num',start_value=-0.6,start_color='F8696B',
    mid_type='num',mid_value=0,mid_color='FFFFFF',
    end_type='num',end_value=0.15,end_color='63BE7B'))

# --- Summary sheet ---
D=np.array([r[10] for r in rows]); ss=wb.create_sheet('Summary')
def hrow(ws,txt):
    ws.append([txt]); ws.cell(ws.max_row,1).font=Font(name=F,bold=True,size=11,color='1F3864')
def trow(ws,vals,bold=False):
    ws.append(vals)
    for j in range(1,len(vals)+1): ss.cell(ss.max_row,j).font=Font(name=F,bold=bold,size=10)
hrow(ss,'COMBINED M1 + refined-v7  vs  CURRENT   (606 players, outfield+RUC, cur>=200)')
ss.append(['Source: RL engine head 8aed420a, store 644d1254; combined = M1 level + refined-v7 band; NOTHING baked (prototype).'])
ss.append(['SCOPE: experienced players only (>=2 seasons). Pre-debut/rookies priced via the pedigree path (v7 not applied) -> unproven value protected per the Harley-Reid guardrail.'])
ss.append([]); trow(ss,['Overall','mean','median','p5','p25','p75','p90','down>2%','flat','up>2%'],True)
trow(ss,['Delta%',round(float(D.mean()),3),round(float(np.median(D)),3),round(float(np.percentile(D,5)),3),
         round(float(np.percentile(D,25)),3),round(float(np.percentile(D,75)),3),round(float(np.percentile(D,90)),3),
         int((D<-.02).sum()),int(((D>=-.02)&(D<=.02)).sum()),int((D>.02).sum())])
for r in range(3,ss.max_row+1):
    for c in [2,3,4,5,6,7]:
        v=ss.cell(r,c).value
        if isinstance(v,float): ss.cell(r,c).number_format='0.0%'
ss.append([]); hrow(ss,'By archetype'); trow(ss,['archetype','n','mean','median','min','max'],True)
for a in ['first-year','rising','plateau','injury-prone','decliner']:
    d=np.array([r[10] for r in rows if r[12]==a])
    trow(ss,[a,len(d)]+([round(float(x),3) for x in [d.mean(),np.median(d),d.min(),d.max()]] if len(d) else ['-','-','-','-']))
ss.append([]); hrow(ss,'By value tier'); trow(ss,['tier','n','mean','median'],True)
for t in ['elite','upper','mid','low']:
    d=np.array([r[10] for r in rows if r[11]==t])
    trow(ss,[t,len(d)]+([round(float(d.mean()),3),round(float(np.median(d)),3)] if len(d) else ['-','-']))
ss.append([]); hrow(ss,'By position'); trow(ss,['pos','n','mean','median','min'],True)
for pos in POS6:
    d=np.array([r[10] for r in rows if r[1]==pos])
    trow(ss,[pos,len(d)]+([round(float(x),3) for x in [d.mean(),np.median(d),d.min()]] if len(d) else ['-','-','-']))
for r in range(1,ss.max_row+1):
    for c in range(3,7):
        v=ss.cell(r,c).value
        if isinstance(v,float): ss.cell(r,c).number_format='0.0%'
ss.column_dimensions['A'].width=16
for c in 'BCDEFGHIJ': ss.column_dimensions[c].width=11

# --- Flags sheet ---
fs=wb.create_sheet('Flags')
def frow(txt,bold=False,col='000000'):
    fs.append([txt]); fs.cell(fs.max_row,1).font=Font(name=F,bold=bold,size=11 if bold else 10,color=col)
frow('FLAG A — trend-up (>=+8) but level-flat (gap<5) established cohort — M1 holds, v7 shaves',True,'1F3864')
frow('Recommendation: trust the LEVEL (walk-forward: trend did NOT separate sustain vs regress, AUC 0.45).',False,'808080')
frow('Apply your reads to the young strongly-trending materially-shaved cases below.',False,'808080')
fs.append([]); fs.append(['Player','Pos','Age','n','Trend','Gap','Current','Delta%'])
for j in range(1,9): fs.cell(fs.max_row,j).font=Font(name=F,bold=True,size=10)
Acohort=[r for r in rows if r[14]=='trend-up/level-flat']; Acohort.sort(key=lambda r:r[10])
for r in Acohort:
    fs.append([r[0],r[1],r[2],r[3],r[6],r[7],r[8],r[10]])
    fs.cell(fs.max_row,8).number_format='0.0%'; fs.cell(fs.max_row,5).number_format='+0.0'; fs.cell(fs.max_row,6).number_format='+0.0'
fs.append([]); frow(f'({len(Acohort)} players)')
fs.append([])
frow('FLAG B — KEY_FWD over-compression (-17.9 mean) — RESOLVED: very likely NOT a bug',True,'1F3864')
frow('Decomposition: these low-value KEY_FWDs have value almost entirely in the upper-band CEILING (q90/q97).',False,'808080')
frow('Both cB (body) and asc (tail) bite the ceiling -> -60% (Cox body-27%+tail-33%; McMahon n0/cB0 all -58% is tail).',False,'808080')
frow('cB-cap barely helps (KEY_FWD -17.9 -> -16.6). It is v7 shaving speculative ceiling value (Ed Richards principle).',False,'808080')
frow('Established valuable key fwds NOT hit (Wright -4.3, Gunston +0). Recommendation: keep asc; accept speculative shaves.',False,'808080')
fs.append([])
frow('SCOPE — v7 applies to EXPERIENCED players (>=2 seasons) only. Pre-debut/rookies are priced via the pedigree path,',False,'808080')
frow('so v7 is NOT applied to them (shaving an unproven pedigree ceiling would violate the Harley-Reid guardrail). The big',False,'808080')
frow('rookie crushes seen when v7 was mis-applied (e.g. Coe pick8, Puncher pick15 — unplayed 2026 draftees) confirm this scope.',False,'808080')
fs.column_dimensions['A'].width=22
for c in 'BCDEFGH': fs.column_dimensions[c].width=11

out='/home/claude/rl_workspace/rl_after/combined_backtest_book_8aed420a.xlsx'
wb.save(out); print("saved",out)
print("sanity top-3 crush:",[ (r[0],round(r[10]*100,1)) for r in rows[:3]])
print("sanity top-3 rise:",[ (r[0],round(r[10]*100,1)) for r in rows[-3:]])
