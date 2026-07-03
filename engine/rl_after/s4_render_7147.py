import os
import json,collections,numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment
from openpyxl.utils import get_column_letter
rec=json.load(open(os.environ.get('S4_MATRIX','s4_matrix.json'))); R=list(rec.values())
TAG=os.environ.get('S4_TAG','8aed420a')   # engine identity shown in sheet titles (display-only)
INCURVE={'ND','RD'}; POOLED=['MSD','SSP','PDA','PDN','IRE','UNR','PDS']
TYPEFULL={'ND':'National','RD':'Rookie','MSD':'Mid-Season','SSP':'SSP','PDA':'PD-Academy','PDN':'PD-NextGen','IRE':'PD-Ireland','UNR':'PD-Unreg','PDS':'PD-Schol'}
for r in R:
    r['anchor']=r['Vpath'][0] if r['Vpath'] else None
    r['cur_disp']=(0 if r.get('retired_now') else r['cur'])      # retired keeper = 0, not the delisted floor
REFONLY={2003,2025}
TITLE=Font(bold=True,size=12);H=Font(bold=True);GREY=PatternFill('solid',fgColor='E8E8E8')
ANCH=PatternFill('solid',fgColor='FFF2CC');REFF=PatternFill('solid',fgColor='FDE9E9');POOLF=PatternFill('solid',fgColor='E2EFDA');BLUE=PatternFill('solid',fgColor='DDEBF7')
wb=Workbook();wb.remove(wb.active)
bycoh=collections.defaultdict(list)
for r in R: bycoh[r['year']].append(r)
cohorts=sorted(k for k in bycoh if k and 2003<=k<=2025)
def num(v): return isinstance(v,(int,float))
GMAX=max((len(x['Vpath']) for x in R if x['type'] in INCURVE and x['year'] in cohorts), default=12)
# ================= SUMMARY (endpoints) =================
sm=wb.create_sheet('Summary')
sm['A1']=f'AFL RL — WALK-FORWARD cohort value BOOK (engine md5 {TAG}, dials + shed + GENTLER upside + NO-GAMES sit-out anchor, 34 positions reconciled to stage0) — FOR EYEBALL before the pick-curve build';sm['A1'].font=TITLE
sm['A2']='V = walk-forward engine value per season (as-of end of that season, only up-to-then data). P = era-adj production. Anchor = end of Year 1 (gold). Current = 0 for retired (worthless keeper), not the delisted floor.'
sm['A3']='CURVE = ND+RD only, 2004-2024 (continuous pick order). 2003 & 2025 = reference-only (excluded). MSD/SSP/PostDraft = STANDALONE pooled (Pooled sheet). See "Cohort Trajectories" for each cohort\'s value season-by-season.'
sm['A4']='POSITIONS reconciled to stage0 (the 34 pre_stage0 discrepancies fixed — Holmes/Sheezel/Rozee/Perryman/etc now authoritative). Pos = DRAFTED pedigree group; switchers show drafted\u2192current. SIT-OUT ANCHOR: still-listed players with 0 seasons of \u22656 games through a year are valued at position-scaled retention \u00d7 their OWN draftval (RUC/KPP gentle, nonKPP steep; plateau yr1-2 then decline). CAVEAT: anchor LEVELS are deferred-to-PVC placeholders and currently over-discount sat-out Yr1 ~16% below the empirical Yr2-realised target, so the Measure-1 peak reads ~158% (absolute peak unchanged; baseline dropped). Nothing baked.'
sm.append([]);sm.append(['Cohort','#ND+RD','#pooled','Yr1 anchor SUM','current SUM (real)','in curve?'])
for c in range(1,7): sm.cell(6,c).font=H; sm.cell(6,c).fill=GREY
for C in cohorts:
    grp=bycoh[C]; ic=[x for x in grp if x['type'] in INCURVE]; pl=[x for x in grp if x['type'] in POOLED]
    asum=sum(x['anchor'] or 0 for x in ic); csum=sum(x['cur_disp'] or 0 for x in ic)
    sm.append([C,len(ic),len(pl),round(asum),round(csum),'YES' if 2004<=C<=2024 else 'reference-only'])
    if C in REFONLY:
        for c in range(1,7): sm.cell(sm.max_row,c).fill=REFF
for col,w in zip('ABCDEF',[10,9,9,16,18,15]): sm.column_dimensions[col].width=w
# ================= COHORT TRAJECTORIES — TWO PERMANENT VIEWS (D8 03/07/2026) =================
# View 1 = FULL HISTORY (curve window 2004-2024, all cohort rows). View 2 = 2015-2024 WINDOW (OVERALL
# row + slopes recomputed over 2015-2024 cohorts only; rows restricted to 2015-2025). Display-only:
# every number is the same walk-forward matrix value — the window changes WHICH cohorts average, nothing else.
def traj_sheet(sheetname,wlo,whi,row_cohorts):
    ct=wb.create_sheet(sheetname)
    ct['A1']=f'COHORT VALUE TRAJECTORIES — {wlo}-{whi} VIEW — each cohort (ND+RD) summed season-by-season since draft. Scan DOWN to compare cohorts, ACROSS to read one cohort build/peak/fade.';ct['A1'].font=TITLE
    ct['A2']='TABLE 1 = MEASURE 1: each year\'s summed V as % of the cohort\'s OWN end-of-Year-1 total. Denominator = FULL COHORT\'s Yr1 value, FIXED across all years (busts INCLUDED forever — they just contribute 0 to the numerator once gone). NO survivor bias, NO PVC.'
    ct['A3']=f'Yr1 = 100% by construction. Later years = value vs the cohort\'s own start -> reads SHAPE (build/peak/fade), NOT cross-cohort quality. Current = 0 for retired. OVERALL row = mean over the {wlo}-{whi} curve cohorts. 2003 & 2025 = reference-only (pink).'
    ct['A4']='MEASURE 2 (cross-cohort quality, vs the UNIVERSAL PVC) is READY TO ADD once the curve exists. TABLE 2 below = absolute summed V (SCAR), reference.'
    curve_cohorts=[C for C in cohorts if wlo<=C<=whi]
    def traj_block(startrow,measure1):
        hdr=['Cohort','#ND+RD','inCurve']+[f'Yr{k}' for k in range(1,GMAX+1)]+['Current']
        for j,h in enumerate(hdr): c=ct.cell(startrow,1+j); c.value=h; c.font=H; c.fill=BLUE
        # OVERALL row = mean across the view's curve cohorts of each Yr column — quantifies the aggregate peak
        percoh={}
        for C in curve_cohorts:
            ic=[x for x in bycoh[C] if x['type'] in INCURVE]; denom=sum(x['anchor'] for x in ic if num(x['anchor']))
            row=[]
            for k in range(1,GMAX+1):
                vals=[x['Vpath'][k-1] for x in ic if len(x['Vpath'])>=k and num(x['Vpath'][k-1])]
                row.append((sum(vals)/denom) if (vals and denom) else None)
            curV=sum(x['cur_disp'] or 0 for x in ic)
            row.append((curV/denom) if denom else None)
            percoh[C]=row
        orow=ct.cell(startrow+1,1); orow.value='OVERALL (mean)'; orow.font=H
        ct.cell(startrow+1,2,len(curve_cohorts)); ct.cell(startrow+1,3,f'{wlo}-{whi}').font=H
        for k in range(GMAX+1):
            col=[percoh[C][k] for C in curve_cohorts if percoh[C][k] is not None]
            if len(col)>=5:
                cc=ct.cell(startrow+1,4+k, float(np.mean(col)) if measure1 else round(float(np.mean(col))*1))
                if measure1: cc.number_format='0%'
                cc.font=H
        for c in range(1,5+GMAX): ct.cell(startrow+1,c).fill=PatternFill('solid',fgColor='FFF6CC')
        rr=startrow+2
        for C in row_cohorts:
            ic=[x for x in bycoh[C] if x['type'] in INCURVE]
            denom=sum(x['anchor'] for x in ic if num(x['anchor']))           # FULL-COHORT end-of-Yr1 total, FIXED across years (busts in)
            ct.cell(rr,1,C); ct.cell(rr,2,len(ic)); ct.cell(rr,3,'YES' if wlo<=C<=whi else 'ref-only')
            for k in range(1,GMAX+1):
                vals=[x['Vpath'][k-1] for x in ic if len(x['Vpath'])>=k and num(x['Vpath'][k-1])]  # survivors that season; busts gone = absent = 0
                if not vals: continue
                if measure1:
                    if denom: cel=ct.cell(rr,3+k,sum(vals)/denom); cel.number_format='0%'
                else: ct.cell(rr,3+k,round(sum(vals)))
            curV=sum(x['cur_disp'] or 0 for x in ic)
            cc=ct.cell(rr,4+GMAX, (curV/denom if (measure1 and denom) else (round(curV) if not measure1 else 0)))
            if measure1: cc.number_format='0%'
            if C in REFONLY or not (wlo<=C<=whi):
                for c in range(1,5+GMAX): ct.cell(rr,c).fill=REFF
            rr+=1
        return rr
    ct.cell(6,1,'TABLE 1 — MEASURE 1: % of cohort\'s OWN end-of-Year-1 total (full cohort, fixed denominator, busts in — Yr1=100%)').font=H
    end1=traj_block(7,True)
    ct.cell(end1+2,1,'TABLE 2 — absolute summed V (SCAR) — reference').font=H
    traj_block(end1+3,False)
    ct.freeze_panes='D8'; ct.column_dimensions['A'].width=8
traj_sheet('Cohort Trajectories',2004,2024,cohorts)                              # VIEW 1 — FULL HISTORY
traj_sheet('Trajectories 2015-2024',2015,2024,[C for C in cohorts if C>=2015])   # VIEW 2 — 2015-2024 window
# ================= PER-COHORT (ND+RD) =================
for C in cohorts:
    ic=sorted([x for x in bycoh[C] if x['type'] in INCURVE], key=lambda x:-(x['cur_disp'] or 0))
    if not ic: continue
    ws=wb.create_sheet(str(C)); ny=max((len(x['Vpath']) for x in ic),default=1)
    tag=' — REFERENCE-ONLY (excluded from curve)' if C in REFONLY else ''
    ws['A1']=f'Cohort {C} — WALK-FORWARD engine VALUE (V) + era-adj PRODUCTION (P), draft→now. Anchor = Yr1 (gold). Current = 0 for retired.{tag}';ws['A1'].font=TITLE
    ws['A2']='Rows sorted by current value. COHORT TOTAL = sum-ratio % of DraftVal. Read SHAPE: V flat while P climbs = option premium realising.'
    hdr=['T','Player','Pos','Type','Pick','DraftVal']+[f'Yr{i}' for i in range(1,ny+1)]+['Current']
    ws.append([]);ws.append(hdr); hr=ws.max_row
    for c in range(1,len(hdr)+1): ws.cell(hr,c).font=H; ws.cell(hr,c).fill=GREY
    first=hr+2
    ws.append(['','COHORT TOTAL (sum-ratio)','','','','']+['']*ny+['']); totr=ws.max_row
    for x in ic:
        V=x['Vpath']+['']*(ny-len(x['Vpath'])); P=x['Ppath']+['']*(ny-len(x['Ppath']))
        ws.append(['V',x['player'][:24],(f"{x['pos']}\u2192{x['cpos']}" if x.get('sw') else x['pos']),TYPEFULL.get(x['type'],x['type']),x['pick'],x['draftval']]+[(round(v) if num(v) else '') for v in V]+[x['cur_disp']])
        ws.cell(ws.max_row,7).fill=ANCH
        ws.append(['P','  ↳ prod','','','','']+[(v if num(v) else '') for v in P]+[''])
    last=ws.max_row
    ws.cell(totr,6).value=f'=SUMIFS(F{first}:F{last},A{first}:A{last},"V")'; ws.cell(totr,6).font=H
    for i in range(ny):
        col=get_column_letter(7+i)
        ws.cell(totr,7+i).value=(f'=IF(SUMIFS(F{first}:F{last},A{first}:A{last},"V",{col}{first}:{col}{last},"<>")=0,"",'
                                 f'SUMIFS({col}{first}:{col}{last},A{first}:A{last},"V")/SUMIFS(F{first}:F{last},A{first}:A{last},"V",{col}{first}:{col}{last},"<>"))')
        ws.cell(totr,7+i).number_format='0%'
    cc=get_column_letter(7+ny); ws.cell(totr,7+ny).value=f'=SUMIFS({cc}{first}:{cc}{last},A{first}:A{last},"V")'; ws.cell(totr,7+ny).font=H
    for c in range(1,len(hdr)+1): ws.cell(totr,c).fill=GREY
    ws.freeze_panes='B'+str(first); ws.column_dimensions['B'].width=22
    for col in ['C','D']: ws.column_dimensions[col].width=11
# ================= POOLED =================
pw=wb.create_sheet('Pooled'); pw['A1']='STANDALONE POOLED mechanisms (NOT in the ND curve). Each → ONE distribution-aware value at the equivalence step. Below = anchor (Yr1 WF) + current per player.';pw['A1'].font=TITLE
pw['A2']='Anchor: MSD ~1.5 seasons; SSP/PostDraft first full-opportunity. Current = 0 for retired. 2024/25-entry pooled players = partial sample (flagged).'
pw.append([]);pw.append(['Mechanism','Player','Pos','EntryYr','Anchor(Yr1 WF)','Current','note'])
for c in range(1,8): pw.cell(4,c).font=H; pw.cell(4,c).fill=GREY
for t in POOLED:
    grp=sorted([x for x in R if x['type']==t], key=lambda x:-(x['cur_disp'] or 0))
    if not grp: continue
    anchors=[x['anchor'] for x in grp if num(x['anchor'])]
    pw.append([f'{TYPEFULL.get(t,t)} ({t}) — n={len(grp)}','','','',round(sum(anchors)/len(anchors)) if anchors else '','mean anchor (ref only; equivalence uses DISTRIBUTION-aware value)',''])
    hr=pw.max_row
    for c in range(1,8): pw.cell(hr,c).fill=POOLF; pw.cell(hr,c).font=H
    for x in grp:
        pw.append(['',x['player'][:24],(f"{x['pos']}\u2192{x['cpos']}" if x.get('sw') else x['pos']),x['year'],x['anchor'],x['cur_disp'],('partial-sample' if (x['year'] or 0)>=2024 else '')])
pw.column_dimensions['B'].width=22; pw.column_dimensions['G'].width=20; pw.column_dimensions['A'].width=22
# ================= FLAT =================
fw=wb.create_sheet('FLAT'); fw.append(['player','pos','type','in_curve','pick','draft_year','season_idx','season_value_WF','is_anchor','production_adj','current_value','retired_now'])
for c in range(1,13): fw.cell(1,c).font=H; fw.cell(1,c).fill=GREY
for x in R:
    for i,(v,p) in enumerate(zip(x['Vpath'],x['Ppath'])):
        fw.append([x['player'],x['pos'],x['type'],('Y' if x['type'] in INCURVE else 'N'),x['pick'],x['year'],i+1,
                   round(v) if num(v) else None,('Y' if i==0 else ''),p,x['cur_disp'],('Y' if x.get('retired_now') else '')])
fw.freeze_panes='A2'
order=['Summary','Cohort Trajectories','Trajectories 2015-2024']+[str(C) for C in cohorts if any(x['type'] in INCURVE for x in bycoh[C])]+['Pooled','FLAT']
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 999)
out=os.environ.get('S4_BOOK','AFL_RL_WALKFORWARD_book.xlsx'); wb.save(out)
print('saved',out,'sheets',len(wb.worksheets))
