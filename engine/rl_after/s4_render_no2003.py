import os, json, collections, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
rec=json.load(open(os.environ.get('S4_MATRIX','s4_matrix_M1v7.json'))); R=list(rec.values())
INCURVE={'ND','RD'}; POOLED={'MSD','SSP','PDA','PDN','IRE','UNR','PDS'}
TYPEFULL={'ND':'National','RD':'Rookie','MSD':'Mid-Season','SSP':'SSP','PDA':'PD-Academy','PDN':'PD-NextGen','IRE':'PD-Ireland','UNR':'PD-Unreg','PDS':'PD-Schol'}
for r in R: r['cur_disp']=(0 if r.get('retired_now') else r['cur'])
REFONLY={2003,2025}
TITLE=Font(bold=True,size=12); H=Font(bold=True); F='Arial'
GREY=PatternFill('solid',fgColor='E8E8E8'); ANCH=PatternFill('solid',fgColor='FFF2CC')
REFF=PatternFill('solid',fgColor='FDE9E9'); BLUE=PatternFill('solid',fgColor='DDEBF7'); HEAD=PatternFill('solid',fgColor='FFF6CC')
POOLF=PatternFill('solid',fgColor='E2EFDA')
def num(v): return isinstance(v,(int,float))
bycoh=collections.defaultdict(list)
for r in R: bycoh[r['year']].append(r)
cohorts=sorted(k for k in bycoh if k and 2004<=k<=2025)  # 2003 fully excluded from this book per request
GMAX=max((len(x['Vpath']) for x in R if x['type'] in INCURVE and x['year'] in cohorts), default=12)
def sumv(pl,k): return sum(x['Vpath'][k-1] for x in pl if len(x['Vpath'])>=k and num(x['Vpath'][k-1]))
def sump(pl,k): return sum(x['Ppath'][k-1] for x in pl if len(x['Ppath'])>=k and num(x['Ppath'][k-1]))
def ratio(pl,k):
    p=sump(pl,k); return (sumv(pl,k)/p) if p>0 else None
wb=Workbook(); wb.remove(wb.active)

# ================= SUMMARY: cohort VALUE indexed to Yr1 anchor = 100% (opportunity-matched) =================
sm=wb.create_sheet('Summary')
sm['A1']='AFL RL — WALK-FORWARD cohort book — M1 + refined-v7 PROTOTYPE (NOT a baked head; engine 8aed420a, store pre_stage0). Nothing baked.'; sm['A1'].font=TITLE
sm['A2']='SUMMARY metric = cohort VALUE indexed to the year-1 anchor = 100% (SUM-RATIO, not mean-of-ratios). For each career-year N: Sum(value at N) / Sum(value at Yr1) x 100. Walk-forward value as-of end of each season (only <=T data; M1+v7 applied).'
sm['A3']='BUSTS = ZERO (no survivor bias): a player who retires stays in the Yr1 denominator but contributes 0 to later years. OPPORTUNITY-MATCHED: each career-year N pools ONLY cohorts old enough to have reached year N (debut+N <= 2026); immature cohorts are excluded from BOTH numerator and denominator for that column (otherwise their Yr1 would deflate later years).'
sm['A4']='HEADLINE = pooled over the opportunity-matched cohorts; the "#cohorts (reached Yr N)" row shows how many contribute to each year (shrinks with N). A SECONDARY "2015-24 pool" line applies the identical opp-matched metric to just the 2015-2024 cohorts (recent-decade sample; tail thins fast — Yr11 is the 2015 cohort only). Each cohort row = its own Sum(value at N)/Sum(value at Yr1), blank once it has not yet reached year N. Pool = ND+RD 2004-2024; 2025 reference-only (pink); 2003 excluded entirely (row + data). Current% = current value as % of Yr1; Current V = absolute current-value sum (retired=0).'
sm['A5']='CAVEAT: sit-out Yr1 anchor LEVELS are deferred-to-PVC placeholders, so absolute % levels are PROVISIONAL until the pick-curve step (read the SHAPE, not the exact %).'
hr=6; hdr=['Cohort','#ND+RD']+[f'Yr{k}' for k in range(1,GMAX+1)]+['Current %','Current V (SCAR)']
for j,h in enumerate(hdr): c=sm.cell(hr,1+j); c.value=h; c.font=H; c.fill=BLUE
def sv1(pl): return sum(x['Vpath'][0] for x in pl if x['Vpath'] and num(x['Vpath'][0]))
def vindex(pool,k):
    elig=[x for x in pool if x['year']+k<=2026]; d=sv1(elig)
    return (sumv(elig,k)/d) if d>0 else None
curve=[x for C in cohorts if 2004<=C<=2024 for x in bycoh[C] if x['type'] in INCURVE]
sm.cell(hr+1,1,'HEADLINE 2004-24 (opp-matched)').font=H; sm.cell(hr+1,2,len(curve)).font=H
for k in range(1,GMAX+1):
    v=vindex(curve,k)
    if v is not None: cc=sm.cell(hr+1,2+k,v); cc.font=H; cc.number_format='0%'
d1=sv1([x for x in curve if x['year']+1<=2026]); curp=(sum(x['cur_disp'] or 0 for x in curve)/d1) if d1>0 else None
if curp is not None: cc=sm.cell(hr+1,2+GMAX+1,curp); cc.font=H; cc.number_format='0%'
sm.cell(hr+1,2+GMAX+2, round(sum(x['cur_disp'] or 0 for x in curve))).font=H
for c in range(1,4+GMAX+1): sm.cell(hr+1,c).fill=HEAD
# --- SECONDARY value-indexed line: 2015-24 cohort sample only (identical opp-matched metric) ---
curve15=[x for C in cohorts if 2015<=C<=2024 for x in bycoh[C] if x['type'] in INCURVE]
sm.cell(hr+2,1,'2015-24 pool (opp-matched)').font=H; sm.cell(hr+2,2,len(curve15)).font=H
for k in range(1,GMAX+1):
    v=vindex(curve15,k)
    if v is not None: cc=sm.cell(hr+2,2+k,v); cc.font=H; cc.number_format='0%'
d1b=sv1([x for x in curve15 if x['year']+1<=2026]); curpb=(sum(x['cur_disp'] or 0 for x in curve15)/d1b) if d1b>0 else None
if curpb is not None: cc=sm.cell(hr+2,2+GMAX+1,curpb); cc.font=H; cc.number_format='0%'
sm.cell(hr+2,2+GMAX+2, round(sum(x['cur_disp'] or 0 for x in curve15))).font=H
for c in range(1,4+GMAX+1): sm.cell(hr+2,c).fill=POOLF
sm.cell(hr+3,1,'#cohorts 2004-24 (reached Yr N)').font=Font(name=F,italic=True,size=9)
for k in range(1,GMAX+1):
    nc=len(set(x['year'] for x in curve if x['year']+k<=2026)); sm.cell(hr+3,2+k,nc).font=Font(name=F,italic=True,size=9)
rr=hr+4
for C in cohorts:
    ic=[x for x in bycoh[C] if x['type'] in INCURVE]
    sm.cell(rr,1,C); sm.cell(rr,2,len(ic)); d1c=sv1(ic)
    for k in range(1,GMAX+1):
        v=vindex(ic,k)
        if v is not None: sm.cell(rr,2+k,v).number_format='0%'
    cp_=(sum(x['cur_disp'] or 0 for x in ic)/d1c) if d1c>0 else None
    if cp_ is not None: sm.cell(rr,2+GMAX+1,cp_).number_format='0%'
    sm.cell(rr,2+GMAX+2, round(sum(x['cur_disp'] or 0 for x in ic)))
    if C in REFONLY:
        for c in range(1,4+GMAX+1): sm.cell(rr,c).fill=REFF
    rr+=1
rr+=1
sm.cell(rr,1,'SECONDARY (not the headline) — pooled \u03a3V/\u03a3P, value per production point').font=Font(name=F,italic=True,size=9)
for k in range(1,GMAX+1):
    v=ratio(curve,k)
    if v is not None: cc=sm.cell(rr,2+k,round(v,1)); cc.font=Font(name=F,italic=True,size=9); cc.number_format='0.0'
sm.freeze_panes='C7'; sm.column_dimensions['A'].width=30; sm.column_dimensions['B'].width=9
for k in range(GMAX+3): sm.column_dimensions[get_column_letter(3+k)].width=7

# ================= PER-COHORT: adjacent V | P columns =================
for C in cohorts:
    ic=sorted([x for x in bycoh[C] if x['type'] in INCURVE], key=lambda x:-(x['cur_disp'] or 0))
    if not ic: continue
    ws=wb.create_sheet(str(C)); ny=max((len(x['Vpath']) for x in ic),default=1)
    tag=' — REFERENCE-ONLY (excluded from curve)' if C in REFONLY else ''
    ws['A1']=f'Cohort {C} — WALK-FORWARD VALUE (V) + era-adj PRODUCTION (P) side-by-side, draft->now (M1+v7 prototype). Yr1 V = anchor (gold). Current = 0 for retired.{tag}'; ws['A1'].font=TITLE
    ws['A2']='Rows = every ND+RD player (busts included, no cherry-pick), sorted by current value. Yr_k V and P are ADJACENT. Read a row ACROSS: V flat while P climbs = option premium realising. Totals up top = Sum(V)/Sum(P) per year.'
    hdr=['Player','Pos','Type','Pick','DraftVal']
    for i in range(1,ny+1): hdr+= [f'Yr{i} V', f'Yr{i} P']
    hdr+=['Current V']
    ws.append([]); ws.append(hdr); hh=ws.max_row
    for c in range(1,len(hdr)+1):
        cell=ws.cell(hh,c); cell.font=H; cell.fill=GREY; cell.alignment=Alignment(horizontal='center')
    # three total rows: ΣV, ΣP, ΣV/ΣP
    rowV=['Σ VALUE','','','','']; rowP=['Σ PRODUCTION','','','','']; rowR=['Σ V / Σ P  (sum-ratio)','','','','']
    for i in range(1,ny+1):
        sv=sumv(ic,i); sp=sump(ic,i)
        rowV += [round(sv),'']; rowP += ['', round(sp,1)]; rowR += [round(sv/sp,1) if sp>0 else '','']
    rowV+=[round(sum(x['cur_disp'] or 0 for x in ic))]; rowP+=['']; rowR+=['']
    for rw,fillbold in [(rowV,True),(rowP,True),(rowR,True)]:
        ws.append(rw)
        for c in range(1,len(hdr)+1):
            cl=ws.cell(ws.max_row,c); cl.font=Font(name=F,bold=True,size=9); cl.fill=HEAD
    # player rows
    for x in ic:
        V=x['Vpath']; P=x['Ppath']
        row=[x['player'][:24],(f"{x['pos']}\u2192{x['cpos']}" if x.get('sw') else x['pos']),TYPEFULL.get(x['type'],x['type']),x['pick'],x['draftval']]
        for i in range(ny):
            row += [ (round(V[i]) if i<len(V) and num(V[i]) else ''), (P[i] if i<len(P) and num(P[i]) else '') ]
        row += [x['cur_disp']]
        ws.append(row)
        for c in range(1,len(hdr)+1): ws.cell(ws.max_row,c).font=Font(name=F,size=9)
        ws.cell(ws.max_row,6).fill=ANCH   # Yr1 V = anchor
        for i in range(ny): ws.cell(ws.max_row,7+2*i).number_format='0.0'  # P columns one decimal
    if C in REFONLY:
        for c in range(1,len(hdr)+1): ws.cell(hh,c).fill=REFF
    ws.freeze_panes='B'+str(hh+1); ws.column_dimensions['A'].width=22
    ws.column_dimensions['B'].width=12; ws.column_dimensions['C'].width=11

# ================= POOLED =================
pw=wb.create_sheet('Pooled'); pw['A1']='STANDALONE POOLED types (NOT in the ND+RD curve): anchor (Yr1 WF value) + current per player. M1+v7 prototype.'; pw['A1'].font=TITLE
pw.append([]); pw.append(['Mechanism','Player','Pos','EntryYr','Anchor Yr1 V','Current V'])
for c in range(1,7): pw.cell(3,c).font=H; pw.cell(3,c).fill=GREY
for t in ['MSD','SSP','PDA','PDN','IRE','UNR','PDS']:
    grp=sorted([x for x in R if x['type']==t], key=lambda x:-(x['cur_disp'] or 0))
    if not grp: continue
    pw.append([f'{TYPEFULL.get(t,t)} ({t}) — n={len(grp)}','','','','',''])
    for c in range(1,7): pw.cell(pw.max_row,c).fill=POOLF; pw.cell(pw.max_row,c).font=H
    for x in grp:
        pw.append(['',x['player'][:24],(f"{x['pos']}\u2192{x['cpos']}" if x.get('sw') else x['pos']),x['year'],
                   (round(x['Vpath'][0]) if x['Vpath'] and num(x['Vpath'][0]) else ''), x['cur_disp']])
pw.column_dimensions['A'].width=22; pw.column_dimensions['B'].width=22

order=['Summary']+[str(C) for C in cohorts if any(x['type'] in INCURVE for x in bycoh[C])]+['Pooled']
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 999)
out=os.environ.get('S4_BOOK','AFL_RL_WALKFORWARD_book_M1v7-proto.xlsx'); wb.save(out)
print('saved',out,'sheets',len(wb.worksheets),'GMAX',GMAX)
print('HEADLINE value indexed to Yr1=100% (pooled 2004-24):',
      [ (f'Yr{k}', (f'{round(vindex(curve,k)*100)}%' if vindex(curve,k) is not None else None)) for k in range(1,min(GMAX,12)+1)])
