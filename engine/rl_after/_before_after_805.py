import io, contextlib, numpy as np
g={}
with contextlib.redirect_stdout(io.StringIO()): exec(open('_merged_recover.py').read().split('print("=== AFTER')[0], g)
MA=g['MA']; cp=g['cp']; ev=g['ev']; delisted=g['delisted']; b6o=g['b6']
_lvlcurr=g['_lvlcurr']; _nqual=g['_nqual']; PROVEN_N=g['PROVEN_N']; DOWN_TOL=g['DOWN_TOL']
_agemult=g['_agemult']; _par_prior=g['_par_prior']; _upS=g['_upS']; _eo=g['_eo']
# ---- the 805: current-listed universe ----
def in805(p): return MA.GRP.get(p.get('pos')) and not delisted(p)
players=[p for p in MA.data if in805(p)]
print("population:",len(players))
MA.BASE_REF=MA.AGE_REF=2026; MA._pe_clear()
# ---- BEFORE (baseline engine value) ----
before={}
for p in players:
    try:
        with contextlib.redirect_stdout(io.StringIO()): before[id(p)]=ev(p,2026)
    except Exception: before[id(p)]=None
# ---- inject M1 + refined-v7 ----
TOL,GA,WIN,S,GC=5.0,12,2,0.46,17.0
def radq(p,Y,Lo): return any(x['games']>=GA and x['avg']>Lo for x in p['scoring'] if Y-WIN<x['year']<=Y and (cp.debutyr(p)-1)<x['year'])
def core(p,Y):
    Lo=cp._lvl_eff_orig(p,Y); n=_nqual(p,Y)
    if n==0: return Lo
    Lc=_lvlcurr(p,Y)
    if n>=PROVEN_N:
        if Lc>=Lo: return (Lo+S*(Lc-Lo)) if ((Lc-Lo)>=TOL and radq(p,Y,Lo)) else Lo
        d=Lo-Lc
        if d<=DOWN_TOL: return Lo
        sw=float(np.clip((d-DOWN_TOL)/5,0,1)); return (1-sw)*Lo+sw*Lc*_agemult(cp._age_asof(p,Y))
    c=n/PROVEN_N; return c*Lc+(1-c)*_par_prior(p,Y)
def infer(p,Y):
    L0=core(p,Y); eo=_eo(p,Y)
    if eo<=0: return L0
    avs=[x['avg'] for x in p['scoring'] if x.get('games',0)>=6 and (cp.debutyr(p)-1)<x['year']<=Y]
    if not avs: return L0
    bar=MA.REPL.get(MA.gfut(p),0.0)-3.0; N=Y-cp.debutyr(p)+1
    return (1-eo)*L0+eo*min(L0,max(_upS(max(avs)-bar,N),_lvlcurr(p,Y)))
def effs(p,Y): return sum(min(x['games']/GC,1.0) for x in p['scoring'] if x['games']>=6 and (cp.debutyr(p)-1)<x['year']<=Y)
def v7(bb,p,Y):
    bb=list(bb); m=bb[2]; a=cp._age_asof(p,Y); cB=0.47*float(np.clip((effs(p,Y)-1)/3,0,1))
    asc=float(np.interp(a,[20,22,24,27],[1.0,0.76,0.58,0.40])); bb[3]=m+(1-cB)*(bb[3]-m); bb[4]=m+(1-cB)*(bb[4]-m); bb[5]=m+asc*(bb[5]-m); return bb
REAL=set(id(p) for p in players)
def b6f(p,Y=2026):
    bb=b6o(p,Y)
    return v7(bb,p,Y) if id(p) in REAL else bb
cp._lvl_eff=infer; g['b6']=b6f
MA._pe_clear()
# ---- AFTER (M1+v7 engine value) ----
after={}
for p in players:
    try:
        with contextlib.redirect_stdout(io.StringIO()): after[id(p)]=ev(p,2026)
    except Exception: after[id(p)]=None
rows=[]
for p in players:
    bf=before[id(p)]; af=after[id(p)]
    dpct=((af-bf)/bf) if (bf and af is not None and bf>0) else None
    pos=(f"{MA.GRP.get(p.get('pos'))}\u2192{MA.gfut(p)}" if (MA.GRP.get(p.get('pos')) and MA.GRP.get(p.get('pos'))!=MA.gfut(p)) else (MA.GRP.get(p.get('pos')) or MA.gfut(p)))
    rows.append([p['player'], p.get('year'), pos, (MA.effpk(p) if not p.get('_pickless') else None), p.get('type'),
                 bf, af, (af-bf if (bf is not None and af is not None) else None), dpct])
rows.sort(key=lambda r:-(r[5] or 0))
print("rows:",len(rows)," dup names:",len(rows)-len(set(r[0] for r in rows)))
D=[r[8] for r in rows if r[8] is not None]
print(f"mean Δ {np.mean([r[7] for r in rows if r[7] is not None]):+.0f} SCAR ; mean Δ% {100*np.mean(D):+.1f}% ; down:{sum(x<0 for x in D)} flat:{sum(x==0 for x in D)} up:{sum(x>0 for x in D)}")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
wb=Workbook(); ws=wb.active; ws.title='Before vs After'; F='Arial'
ws['A1']='AFL RL — BEFORE vs AFTER engine value, all current players (n=805). BEFORE = baseline (head 8aed420a). AFTER = M1 + refined-v7 PROTOTYPE. Nothing baked. Value = SCAR, as-of 2026.'
ws['A1'].font=Font(name=F,bold=True,size=11); ws.merge_cells('A1:I1')
cols=['Player','DraftYr','Pos','Pick','Type','Before','After','Δ','Δ%']
ws.append([]); ws.append(cols); hr=ws.max_row
hf=PatternFill('solid',fgColor='1F3864')
for j,c in enumerate(cols,1):
    cell=ws.cell(hr,j); cell.font=Font(name=F,bold=True,color='FFFFFF'); cell.fill=hf; cell.alignment=Alignment(horizontal='center')
for r in rows: ws.append(r)
n=ws.max_row; thin=Side(style='thin',color='E0E0E0')
for i in range(hr+1,n+1):
    for j in range(1,10): ws.cell(i,j).font=Font(name=F,size=9); ws.cell(i,j).border=Border(bottom=thin)
    ws.cell(i,6).number_format='#,##0'; ws.cell(i,7).number_format='#,##0'; ws.cell(i,8).number_format='+#,##0;-#,##0'; ws.cell(i,9).number_format='+0.0%;-0.0%'
widths=[24,8,14,6,12,9,9,9,8]
for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w
ws.freeze_panes=f'A{hr+1}'; ws.auto_filter.ref=f'A{hr}:I{n}'
ws.conditional_formatting.add(f'I{hr+1}:I{n}', ColorScaleRule(start_type='num',start_value=-0.6,start_color='F8696B',mid_type='num',mid_value=0,mid_color='FFFFFF',end_type='num',end_value=0.15,end_color='63BE7B'))
out='/home/claude/rl_workspace/rl_after/AFL_RL_before_after_M1v7_805.xlsx'; wb.save(out)
print("saved",out)
print("top5 by value:",[(r[0],r[5],r[6]) for r in rows[:5]])
print("biggest UP:",sorted([r for r in rows if r[8] is not None],key=lambda r:-r[8])[:3])
print("biggest DOWN:",sorted([r for r in rows if r[8] is not None],key=lambda r:r[8])[:3])
