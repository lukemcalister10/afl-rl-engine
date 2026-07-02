"""Cohort backtest book: tracks each draft cohort's collective value vs the PVC they were drafted at,
year by year (engine value() as-of each season via BASE_REF), plus a current-dist snapshot and a
non-play empirical analysis. Run from rl_after:  PYTHONHASHSEED=0 python3 ../forward_valuation/build_cohort_book.py"""
import os,sys,io,contextlib
os.environ.setdefault('RL_GAMMA','0.85'); os.environ.setdefault('RL_PICK1','3000')
sys.path.insert(0,'/home/claude/rl_after')
with contextlib.redirect_stdout(io.StringIO()): import rl_model as MA
import importlib.util,numpy as np
spec=importlib.util.spec_from_file_location('dp','/home/claude/rl_workspace/forward_valuation/distribution_pricing.py')
dp=importlib.util.module_from_spec(spec)
with contextlib.redirect_stdout(io.StringIO()): spec.loader.exec_module(dp)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NOW=2026
COHorts=list(range(2009,2026))                      # 2009..2025
def drafted(C):                                     # cohort members: drafted year C, real group, real pick, not a credit phantom
    out=[p for p in MA.data if p.get('year')==C and not p.get('_double_count') and MA.GRP.get(p['pos'])]
    # collapse duplicate-key (traded-club) rows to one, prefer fuller history
    best={}
    for p in out:
        k=(p['key'] or MA.slug(p['player']))
        if k not in best or len(p['scoring'])>len(best[k]['scoring']): best[k]=p
    return sorted(best.values(), key=lambda p:MA.effpk(p))
def pvc(p): return MA.PVC[min(MA.effpk(p),70)]

# --- as-of value() trajectory: loop years, set BASE_REF once per year, value all relevant players ---
COH={C:drafted(C) for C in COHorts}
allcoh=[(C,p) for C in COHorts for p in COH[C]]
print(f'cohorts {COHorts[0]}-{COHorts[-1]}, {len(allcoh)} drafted players total',file=sys.stderr)
ASOF={}                                             # (id(p), Y) -> value() as the engine WOULD HAVE seen it at year Y (walk-forward)
for Y in range(2009, NOW+1):
    # WALK-FORWARD: temporarily truncate every player's scoring to rows <= Y, so debut_factor / peak_est / los_decay /
    # level_now ALL respect the as-of year. Without this, debut_factor & peak_est read the whole career record and leak
    # future games into a past valuation (e.g. a 2025-drafted prospect's 2026 debut inflating his as-of-2025 value).
    saved={}
    for C,p in allcoh:
        if C>Y: continue
        saved[id(p)]=p['scoring']; p['scoring']=[r for r in p['scoring'] if r['year']<=Y]
    MA.BASE_REF=Y; MA.AGE_REF=Y; MA._pe_clear()
    for C,p in allcoh:
        if C>Y: continue
        try: ASOF[(id(p),Y)]=MA.value(p)
        except Exception: ASOF[(id(p),Y)]=None
    for C,p in allcoh:                               # restore full scoring before the next year / before dist + non-play
        if id(p) in saved: p['scoring']=saved[id(p)]
    MA._pe_clear()
MA.BASE_REF=MA.AGE_REF=NOW; MA._pe_clear()
print('as-of trajectory computed',file=sys.stderr)

# --- current dist snapshot ---
models,prior,META=dp.build()
DIST={}
for C,p in allcoh:
    try: DIST[id(p)]=dp.dist_value(p,models,prior,META,30.0,1.0)
    except Exception: DIST[id(p)]=None
print('current dist computed',file=sys.stderr)

# ---------- workbook ----------
wb=Workbook(); HDR=Font(name='Arial',bold=True,color='FFFFFF'); BOLD=Font(name='Arial',bold=True)
NORM=Font(name='Arial'); HFILL=PatternFill('solid',fgColor='1F3864'); SUBFILL=PatternFill('solid',fgColor='D9E1F2')
TOTFILL=PatternFill('solid',fgColor='FFF2CC'); CEN=Alignment(horizontal='center'); THIN=Side(style='thin',color='BFBFBF')
BORD=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
def style_hdr(ws,row,ncol):
    for c in range(1,ncol+1):
        cell=ws.cell(row,c); cell.font=HDR; cell.fill=HFILL; cell.alignment=CEN; cell.border=BORD
PCT='0%'

# ===== SUMMARY =====
ws=wb.active; ws.title='Summary'
ws['A1']='COHORT BACKTEST — collective value vs PVC drafted at'; ws['A1'].font=Font(name='Arial',bold=True,size=14)
ws['A2']=('Each row = a draft cohort. yr0..yr10 = the cohort\'s combined engine value() AS OF that many years after the draft, '
          'as a % of the total PVC the cohort was drafted at (BASE_REF as-of, walk-forward). "Now dist" = current distribution-model '
          'value vs PVC. Lower-triangular: blank = that year is in the future for that cohort.')
ws['A2'].font=Font(name='Arial',italic=True,size=9); ws['A2'].alignment=Alignment(wrap_text=True,vertical='top'); ws.row_dimensions[2].height=42
ws.merge_cells('A2:R2')
hrow=4
maxk=10
heads=['Cohort','n','PVC drafted']+[f'yr{k}' for k in range(maxk+1)]+['Now value()','Now dist']
for j,h in enumerate(heads,1): ws.cell(hrow,j,h)
style_hdr(ws,hrow,len(heads))
r=hrow+1
sumrows={}
for C in COHorts:
    coh=COH[C]; n=len(coh); PV=sum(pvc(p) for p in coh)
    ws.cell(r,1,str(C)).font=BOLD; ws.cell(r,2,n).font=NORM; ws.cell(r,3,PV).font=NORM; ws.cell(r,3).number_format='#,##0'
    for k in range(maxk+1):
        Y=C+k
        if Y>NOW: continue
        sv=sum(v for p in coh if (v:=ASOF.get((id(p),Y))) is not None)
        cell=ws.cell(r,4+k, sv/PV if PV else None); cell.number_format=PCT; cell.font=NORM
    # now value() (==yr at NOW) and now dist
    nowv=sum(v for p in coh if (v:=ASOF.get((id(p),NOW))) is not None)
    nowd=sum(v for p in coh if (v:=DIST.get(id(p))) is not None)
    ws.cell(r,4+maxk+1, nowv/PV if PV else None).number_format=PCT
    ws.cell(r,4+maxk+2, nowd/PV if PV else None).number_format=PCT
    ws.cell(r,4+maxk+2).font=BOLD
    sumrows[C]=r; r+=1
# col widths
ws.column_dimensions['A'].width=8; ws.column_dimensions['B'].width=5; ws.column_dimensions['C'].width=12
for k in range(maxk+1): ws.column_dimensions[get_column_letter(4+k)].width=6
ws.column_dimensions[get_column_letter(4+maxk+1)].width=11; ws.column_dimensions[get_column_letter(4+maxk+2)].width=9
ws.freeze_panes='D5'

# ===== PER-COHORT TABS =====
for C in COHorts:
    coh=COH[C]
    if len(coh)<10: continue
    ws=wb.create_sheet(str(C))
    ws['A1']=f'{C} DRAFT COHORT — value() as-of each year vs PVC drafted at'; ws['A1'].font=Font(name='Arial',bold=True,size=12)
    years=list(range(C,NOW+1))
    heads=['Player','Pos','Pick','PVC']+[str(Y) for Y in years]+['Now dist','Now ÷PVC']
    hrow=3
    for j,h in enumerate(heads,1): ws.cell(hrow,j,h)
    style_hdr(ws,hrow,len(heads))
    r=hrow+1; first=r
    pvc_col=4
    for p in coh:
        ws.cell(r,1,p['player']).font=NORM
        ws.cell(r,2,MA.gfut(p)).font=NORM
        ws.cell(r,3,MA.effpk(p)).font=NORM
        ws.cell(r,4,pvc(p)).font=NORM; ws.cell(r,4).number_format='#,##0'
        for jY,Y in enumerate(years):
            v=ASOF.get((id(p),Y)); cell=ws.cell(r,5+jY, v if v is not None else None)
            cell.number_format='#,##0'; cell.font=NORM
        dcol=5+len(years)
        d=DIST.get(id(p)); ws.cell(r,dcol, d if d is not None else None).number_format='#,##0'; ws.cell(r,dcol).font=BOLD
        # ratio formula  Now dist / PVC
        ws.cell(r,dcol+1, f'={get_column_letter(dcol)}{r}/{get_column_letter(pvc_col)}{r}').number_format='0%'
        r+=1
    last=r-1
    # totals row (formulas)
    ws.cell(r,1,'COHORT TOTAL').font=BOLD; ws.cell(r,1).fill=TOTFILL
    ws.cell(r,4,f'=SUM({get_column_letter(4)}{first}:{get_column_letter(4)}{last})').font=BOLD
    ws.cell(r,4).number_format='#,##0'; ws.cell(r,4).fill=TOTFILL
    for jY,Y in enumerate(years):
        col=get_column_letter(5+jY)
        ws.cell(r,5+jY,f'=SUM({col}{first}:{col}{last})').font=BOLD; ws.cell(r,5+jY).number_format='#,##0'; ws.cell(r,5+jY).fill=TOTFILL
    dcol=5+len(years); dl=get_column_letter(dcol)
    ws.cell(r,dcol,f'=SUM({dl}{first}:{dl}{last})').font=BOLD; ws.cell(r,dcol).number_format='#,##0'; ws.cell(r,dcol).fill=TOTFILL
    # retention row: total / PVC total
    r2=r+1; ws.cell(r2,1,'÷ PVC').font=BOLD; ws.cell(r2,1).fill=TOTFILL
    pl=get_column_letter(4)
    for jY,Y in enumerate(years):
        col=get_column_letter(5+jY)
        ws.cell(r2,5+jY,f'={col}{r}/{pl}{r}').font=BOLD; ws.cell(r2,5+jY).number_format='0%'; ws.cell(r2,5+jY).fill=TOTFILL
    ws.cell(r2,dcol,f'={dl}{r}/{pl}{r}').font=BOLD; ws.cell(r2,dcol).number_format='0%'; ws.cell(r2,dcol).fill=TOTFILL
    ws.column_dimensions['A'].width=22; ws.column_dimensions['B'].width=9; ws.column_dimensions['C'].width=6; ws.column_dimensions['D'].width=8
    for jY in range(len(years)): ws.column_dimensions[get_column_letter(5+jY)].width=7
    ws.column_dimensions[get_column_letter(dcol)].width=9; ws.column_dimensions[get_column_letter(dcol+1)].width=8
    ws.freeze_panes='E4'

# ===== NON-PLAY ANALYSIS =====
ws=wb.create_sheet('Non-play analysis')
ws['A1']='NON-PLAY EMPIRICAL ANALYSIS — what actually happens to drafted players by how soon they play'; ws['A1'].font=Font(name='Arial',bold=True,size=12)
ws['A2']=('Resolved cohorts (drafted 2010-2020, >=5 seasons of data). For each player: did they play in year 1 (draft+1) and year 2? '
          'Eventual = REALIZED production value (career best-3 season average run through the production chain, NO pedigree prop - so a '
          'bust shows as a bust). Shown vs PVC drafted at. This calibrates how much an unplayed prospect should be discounted, by position and pick band. '
          'mean val÷PVC and hit-rate now reflect real outcomes, not pedigree.')
ws['A2'].font=Font(name='Arial',italic=True,size=9); ws['A2'].alignment=Alignment(wrap_text=True,vertical='top'); ws.row_dimensions[2].height=42; ws.merge_cells('A2:H2')

RES=[p for C in range(2010,2021) for p in COH[C]]
def played(p,k):  # played in year draft+k ?
    return any(x['year']==p['year']+k and x['games']>=1 for x in p['scoring'])
def cb3(p):  # realized career best-3 season average (games>=6) — pure ground-truth production
    a=sorted([x['avg'] for x in p['scoring'] if x['games']>=6],reverse=True)[:3]
    return float(np.mean(a)) if a else None
def evt(p):  # eventual REALIZED production value (no pedigree): value of a player whose peak = their actual best-3
    c=cb3(p)
    if c is None: return 0.0                         # never established a season -> ~0 realized value (a real bust)
    g=MA.gfut(p)
    return MA.val(MA.proj_from_peak(g,c,MA.PEAK_AGE[g],c,'bal'))
def band(pk): return '1-10' if pk<=10 else '11-20' if pk<=20 else '21-40' if pk<=40 else '41+'
def grp(p): return MA.gfut(p)

def summarise(rows):
    rows=[p for p in rows if evt(p) is not None and pvc(p)]
    if not rows: return (0,None,None,None,None)
    rets=[evt(p)/pvc(p) for p in rows]
    hit=sum(1 for p in rows if evt(p)>=pvc(p))/len(rows)
    return (len(rows), float(np.mean([evt(p) for p in rows])), float(np.mean([pvc(p) for p in rows])),
            float(np.median(rets)), hit)

# Block 1: by year-1 play status x pick band
hrow=4; ws.cell(hrow,1,'BY YEAR-1 PLAY STATUS x PICK BAND').font=BOLD
heads=['Played yr1?','Pick band','n','mean eventual val','mean PVC','median val÷PVC','hit-rate (val>=PVC)']
hrow+=1
for j,h in enumerate(heads,1): ws.cell(hrow,j,h)
style_hdr(ws,hrow,len(heads)); r=hrow+1
for pl in [True,False]:
    for bd in ['1-10','11-20','21-40','41+']:
        grp_rows=[p for p in RES if played(p,1)==pl and band(MA.effpk(p))==bd]
        n,mv,mp,mr,hr=summarise(grp_rows)
        if n==0: continue
        ws.cell(r,1,'Yes' if pl else 'No').font=NORM; ws.cell(r,2,bd).font=NORM; ws.cell(r,3,n).font=NORM
        ws.cell(r,4,round(mv)).number_format='#,##0'; ws.cell(r,5,round(mp)).number_format='#,##0'
        ws.cell(r,6,mr).number_format='0%'; ws.cell(r,7,hr).number_format='0%'
        if not pl: 
            for c in range(1,8): ws.cell(r,c).fill=SUBFILL
        r+=1

# Block 2: by year-1 play status x position
r+=1; ws.cell(r,1,'BY YEAR-1 PLAY STATUS x POSITION').font=BOLD; r+=1
for j,h in enumerate(['Played yr1?','Position','n','mean eventual val','mean PVC','median val÷PVC','hit-rate'],1): ws.cell(r,j,h)
style_hdr(ws,r,7); r+=1
for pl in [True,False]:
    for g in ['MID','GEN_DEF','GEN_FWD','KEY_DEF','KEY_FWD','RUC']:
        grp_rows=[p for p in RES if played(p,1)==pl and grp(p)==g]
        n,mv,mp,mr,hr=summarise(grp_rows)
        if n==0: continue
        ws.cell(r,1,'Yes' if pl else 'No').font=NORM; ws.cell(r,2,g).font=NORM; ws.cell(r,3,n).font=NORM
        ws.cell(r,4,round(mv)).number_format='#,##0'; ws.cell(r,5,round(mp)).number_format='#,##0'
        ws.cell(r,6,mr).number_format='0%'; ws.cell(r,7,hr).number_format='0%'
        if not pl:
            for c in range(1,8): ws.cell(r,c).fill=SUBFILL
        r+=1

# Block 3: still-unplayed-by-yr2 vs played-by-yr2 (the Smillie/Patterson case), overall + by band
r+=1; ws.cell(r,1,'NEVER PLAYED THROUGH YEAR 2 vs PLAYED BY YEAR 2 (the compounding case)').font=BOLD; r+=1
for j,h in enumerate(['Status','Pick band','n','mean eventual val','mean PVC','median val÷PVC','hit-rate'],1): ws.cell(r,j,h)
style_hdr(ws,r,7); r+=1
def played_by2(p): return played(p,1) or played(p,2)
for st,fn in [('played by yr2',lambda p:played_by2(p)),('NOT played thru yr2',lambda p:not played_by2(p))]:
    for bd in ['1-10','11-20','21-40','41+']:
        grp_rows=[p for p in RES if fn(p) and band(MA.effpk(p))==bd]
        n,mv,mp,mr,hr=summarise(grp_rows)
        if n==0: continue
        ws.cell(r,1,st).font=NORM; ws.cell(r,2,bd).font=NORM; ws.cell(r,3,n).font=NORM
        ws.cell(r,4,round(mv)).number_format='#,##0'; ws.cell(r,5,round(mp)).number_format='#,##0'
        ws.cell(r,6,mr).number_format='0%'; ws.cell(r,7,hr).number_format='0%'
        if 'NOT' in st:
            for c in range(1,8): ws.cell(r,c).fill=SUBFILL
        r+=1
for col,w in [('A',18),('B',10),('C',5),('D',18),('E',12),('F',16),('G',18)]: ws.column_dimensions[col].width=w

out='/mnt/user-data/outputs/AFL_cohort_backtest_book_2026.xlsx'
wb.save(out); print('saved',out)
